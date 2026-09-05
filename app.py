"""
FinSight AI
===========
A Flask web application for cleaning, analysing, and forecasting
financial data with company-based data isolation and a per-company
Power BI dashboard link.

Entry point:  python app.py
Database init: python init_db.py
"""

from __future__ import annotations

import io
import json
import mimetypes
import os
import re
import secrets
import shutil
import tempfile
import threading
import traceback
import uuid
from datetime import datetime, date, timedelta
from functools import wraps
from pathlib import Path
from email.utils import parseaddr
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

# Load local development variables without overriding deployment variables.
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
except Exception:
    pass

import base64
import matplotlib
matplotlib.use("Agg")  # non-interactive backend
import matplotlib.pyplot as plt
import mysql.connector
from mysql.connector import pooling, errorcode
import numpy as np
import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill
from flask import (
    Flask, render_template, request, jsonify, redirect, url_for,
    session, flash, send_file,
)
try:
    from sklearn.linear_model import LinearRegression
    from sklearn.model_selection import train_test_split
    from sklearn.metrics import mean_absolute_error, r2_score
except ImportError as exc:
    # Keep health checks, authentication, uploads, analytics, and chat
    # available if an optional native ML wheel is unavailable on a host.
    print(f"[finsight] Warning: ML dependencies unavailable: {type(exc).__name__}")
    LinearRegression = None
    train_test_split = None
    mean_absolute_error = None
    r2_score = None
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from data_mapping import (
    CANONICAL_NUMERIC_FIELDS,
    CANONICAL_TEXT_FIELDS,
    FIELD_SPECS,
    apply_mapping,
    clean_dataframe,
    detect_columns,
    json_safe_record,
    profile_dataframe,
    profile_json_payload,
)

# ML Pipeline - improved ML with proper validation and no leakage
try:
    from ml_pipeline import get_pipeline, RuleBasedRiskModel
except ImportError as e:
    print(f"[finsight] Warning: Could not import ml_pipeline: {e}")
    get_pipeline = None
    RuleBasedRiskModel = None

# Universal analysis - dataset-agnostic cleaning / target detection / models
try:
    import universal_analysis
except ImportError as e:
    print(f"[finsight] Warning: Could not import universal_analysis: {e}")
    universal_analysis = None

try:
    import i18n
except ImportError as e:
    print(f"[finsight] Warning: Could not import i18n: {e}")
    i18n = None

from init_db import create_tables as create_database_tables


# =====================================================
# Configuration
# =====================================================
BASE_DIR = Path(__file__).resolve().parent


def _env(name: str, default: str = "") -> str:
    return os.environ.get(name, default).strip()


_DB_ENV_ALIASES = {
    "DB_HOST": ("DB_HOST", "MYSQL_HOST"),
    "DB_PORT": ("DB_PORT", "MYSQL_PORT"),
    "DB_USER": ("DB_USER", "MYSQL_USER"),
    "DB_PASSWORD": ("DB_PASSWORD", "MYSQL_PASSWORD"),
    "DB_NAME": ("DB_NAME", "MYSQL_DATABASE"),
    "DB_SSL_CA": ("DB_SSL_CA", "MYSQL_SSL_CA"),
}


def _db_env(name: str, default: str = "") -> str:
    """Read the documented DB_* names and legacy MYSQL_* aliases."""
    for candidate in _DB_ENV_ALIASES.get(name, (name,)):
        value = _env(candidate)
        if value:
            return value
    return default


def _int_env(name: str, default: int) -> int:
    try:
        return int(_env(name, str(default)))
    except Exception:
        return default


def _bool_env(name: str, default: bool) -> bool:
    return _env(name, str(default)).lower() in {"1", "true", "yes", "on"}


_db_ssl_ca_lock = threading.Lock()
_db_ssl_ca_source: str | None = None
_db_ssl_ca_path: str | None = None


def _ssl_ca_path(value: str) -> str:
    """Return a connector-readable CA path for a path or PEM environment value."""
    if not value.startswith("-----BEGIN CERTIFICATE-----"):
        return value

    pem = value.replace("\\r\\n", "\n").replace("\\n", "\n").replace("\r\n", "\n").replace("\r", "\n")
    global _db_ssl_ca_source, _db_ssl_ca_path
    with _db_ssl_ca_lock:
        if _db_ssl_ca_source == pem and _db_ssl_ca_path and os.path.isfile(_db_ssl_ca_path):
            return _db_ssl_ca_path

        fd, path = tempfile.mkstemp(prefix="finsight-aiven-", suffix=".pem")
        try:
            with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as cert_file:
                cert_file.write(pem)
                if not pem.endswith("\n"):
                    cert_file.write("\n")
            os.chmod(path, 0o600)
        except Exception:
            try:
                os.close(fd)
            except OSError:
                pass
            try:
                os.unlink(path)
            except OSError:
                pass
            raise

        _db_ssl_ca_source = pem
        _db_ssl_ca_path = path
        return path


def _db_config() -> dict[str, object]:
    """Build verified TLS configuration, materializing PEM CA values when needed."""
    required = ("DB_HOST", "DB_PORT", "DB_USER", "DB_NAME")
    missing = [name for name in required if not _db_env(name)]
    if missing:
        raise RuntimeError("Missing required database environment variable(s): " + ", ".join(missing))
    try:
        port = int(_db_env("DB_PORT"))
    except ValueError as exc:
        raise RuntimeError("DB_PORT must be a number") from exc
    if not 1 <= port <= 65535:
        raise RuntimeError("DB_PORT must be between 1 and 65535")
    config = {
        "host": _db_env("DB_HOST"),
        "port": port,
        "user": _db_env("DB_USER"),
        "password": _db_env("DB_PASSWORD"),
        "database": _db_env("DB_NAME"),
        "ssl_verify_cert": True,
        "ssl_verify_identity": True,
    }
    ssl_ca = _db_env("DB_SSL_CA")
    if ssl_ca:
        config["ssl_ca"] = _ssl_ca_path(ssl_ca)
    return config


SECRET_KEY = _env("SECRET_KEY") or secrets.token_hex(32)
UPLOAD_FOLDER = BASE_DIR / _env("UPLOAD_FOLDER", "uploads")
UPLOAD_FOLDER.mkdir(exist_ok=True)
POWERBI_ROOT = BASE_DIR / _env("POWERBI_ROOT", "users")
POWERBI_ROOT.mkdir(exist_ok=True)
POWERBI_TEMPLATE = BASE_DIR / _env("POWERBI_TEMPLATE", "finsightai.pbix")
MAX_CONTENT_LENGTH = _int_env("MAX_CONTENT_LENGTH", 16 * 1024 * 1024)
ALLOWED_EXTENSIONS = {
    e.lower() for e in _env("ALLOWED_EXTENSIONS", "csv,xlsx,xls,json").split(",") if e
} | {"csv", "xlsx", "xls", "json"}
MAX_DATA_COLUMNS = _int_env("MAX_DATA_COLUMNS", 200)
OPENROUTER_API_KEY = _env("OPENROUTER_API_KEY")
OPENROUTER_MODEL = _env("OPENROUTER_MODEL", "openrouter/free")
GROQ_TIMEOUT = max(10, min(120, _int_env("GROQ_TIMEOUT", 45)))




app = Flask(__name__)
_production_mode = _env("FLASK_ENV", "production").lower() == "production"
app.config.update(
    SECRET_KEY=SECRET_KEY,
    MAX_CONTENT_LENGTH=MAX_CONTENT_LENGTH,
    UPLOAD_FOLDER=str(UPLOAD_FOLDER),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=_bool_env("SESSION_COOKIE_SECURE", _production_mode),
)
if _production_mode and not _env("SECRET_KEY"):
    app.logger.warning("SECRET_KEY is not configured; sessions will be invalidated when the server restarts.")
if i18n is not None:
    i18n.init_app(app)


# =====================================================
# MySQL connection pool (lazy - created on first request so that
# any updates to .env credentials are always picked up).
# =====================================================
db_pool = None


def _build_pool():
    global db_pool
    if db_pool is not None:
        return db_pool
    try:
        db_pool = pooling.MySQLConnectionPool(
            pool_name="finsight_pool",
            pool_size=5,
            pool_reset_session=True,
            **_db_config(),
        )
        print("[finsight] MySQL connection pool ready.")
    except Exception as exc:
        print(f"[finsight] WARNING: could not create pool ({exc}). Falling back to direct connections.")
        db_pool = None
    return db_pool


def get_db():
    """Return a fresh MySQL connection (uses pool when available)."""
    pool = _build_pool()
    if pool is not None:
        try:
            return pool.get_connection()
        except mysql.connector.Error as exc:
            # If the cached pool contains a stale connection (wrong
            # credentials that were valid at startup, or a closed
            # socket), rebuild the pool once and retry.
            print(f"[finsight] Pool get_connection failed ({exc}); rebuilding pool.")
            global db_pool
            try:
                db_pool = None
                pool = _build_pool()
                if pool is not None:
                    return pool.get_connection()
            except Exception as rebuild_exc:
                # Keep the original connector failure visible in the log; a
                # silent fallback makes a bad deployment configuration hard to
                # distinguish from a stale pooled connection.
                print(
                    "[finsight] Pool rebuild failed "
                    f"({type(rebuild_exc).__name__}: {rebuild_exc})."
                )
            # Last resort - direct connect with the current environment configuration.
            return mysql.connector.connect(**_db_config())
    return mysql.connector.connect(**_db_config())


def run_query(sql: str, params: tuple | list | None = None, *, fetchone=False, fetchall=False, commit=False):
    """Convenience helper for queries that don't need a long-lived cursor."""
    conn = get_db()
    cur = None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(sql, params or ())
        result = None
        if fetchone:
            result = cur.fetchone()
        elif fetchall:
            result = cur.fetchall()
        if commit:
            conn.commit()
        last_id = cur.lastrowid
        return {"row": result, "rows": result if fetchall else None, "last_id": last_id}
    except Exception:
        if commit:
            try:
                conn.rollback()
            except Exception:
                pass
        raise
    finally:
        if cur is not None:
            cur.close()
        conn.close()


# Existing deployments may use the older users-table naming convention. Keep
# the aliases deliberately small and only interpolate names returned by the
# database itself into SQL identifiers.
_USER_COLUMN_ALIASES = {
    "display": ("name", "full_name", "username", "user_name", "display_name"),
    "first_name": ("firstName", "first_name", "firstname"),
    "last_name": ("lastName", "last_name", "lastname"),
    "email": ("email", "email_address"),
    "password": ("password_hash", "password", "passwd", "hashed_password"),
    "company_id": ("company_id", "companyId"),
    "role": ("role", "user_role"),
    "created_at": ("created_at", "createdAt"),
}


def _quote_identifier(value: str) -> str:
    return "`" + value.replace("`", "``") + "`"


def _find_user_columns(cursor) -> dict[str, str]:
    """Discover live users-table columns without querying information_schema.

    Some hosted MySQL accounts can query ``users`` but cannot see its
    ``information_schema`` rows. A zero-row SELECT still exposes the table's
    column metadata through the connector and works for both old and current
    schemas.
    """
    cursor.execute("SELECT * FROM users LIMIT 0")
    cursor.fetchall()
    available = {}
    for metadata in cursor.description or ():
        value = metadata[0] if metadata else None
        if value:
            available[str(value).lower()] = str(value)

    result = {}
    for logical_name, aliases in _USER_COLUMN_ALIASES.items():
        for alias in aliases:
            actual = available.get(alias.lower())
            if actual:
                result[logical_name] = actual
                break
    return result


def _load_user_columns() -> dict[str, str]:
    """Return column names from the live users table."""
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    try:
        return _find_user_columns(cursor)
    finally:
        cursor.close()
        conn.close()


def _user_display_expression(columns: dict[str, str], qualifier: str = "u") -> str:
    display = columns.get("display")
    if display:
        return f"{qualifier}.{_quote_identifier(display)} AS name"
    first = columns.get("first_name")
    last = columns.get("last_name")
    if first and last:
        return (
            f"CONCAT({qualifier}.{_quote_identifier(first)}, ' ', "
            f"{qualifier}.{_quote_identifier(last)}) AS name"
        )
    if first or last:
        return f"{qualifier}.{_quote_identifier(first or last)} AS name"
    return "NULL AS name"


# =====================================================
# Per-request in-memory data
# =====================================================
# Each user session keeps its own dataframe + models so that companies
# do not share data. The MySQL tables are the durable source of truth.
session_data: dict[int, dict] = {}


def _session_user_id() -> int | None:
    uid = session.get("user_id")
    return int(uid) if uid else None


def _session_company_id() -> int | None:
    cid = session.get("company_id")
    return int(cid) if cid else None


def _message(key: str, **kwargs) -> str:
    """Localized API/flash text; keeps dynamic responses in the active locale."""
    return i18n.t(key, **kwargs) if i18n is not None else key


def _format_number(value: object) -> str:
    """Format a measured value without assigning it a currency meaning."""
    try:
        number = float(value)
        if not np.isfinite(number):
            return "—"
        return f"{number:,.2f}".rstrip("0").rstrip(".")
    except (TypeError, ValueError):
        return str(value) if value is not None else "—"


def _db_text(value: object) -> str | None:
    """Convert a cleaned arbitrary text value to a safe SQL VARCHAR value."""
    if value is None:
        return None
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return str(value)


def _localized_insight(text: str) -> str:
    """Translate deterministic universal-analysis sentences at render time."""
    if not text or i18n is None or i18n.get_locale() == "en":
        return text
    parts = []
    for sentence in text.split(". "):
        match = re.match(r"(.+) show an upward trend, indicating positive growth\.?", sentence)
        if match:
            parts.append(_message("insight.trend.up", names=match.group(1)))
        elif sentence.startswith("The model points to the main business indicators"):
            parts.append(_message("insight.model_strong"))
        elif sentence.startswith("The most relevant outcome is "):
            label = sentence.split('"')[1] if '"' in sentence else ""
            parts.append(_message("insight.classification.focus", pred=label))
        else:
            parts.append(sentence)
    return " ".join(parts)


def get_models(user_id: int) -> dict:
    return session_data.setdefault(user_id, {}).setdefault("models", {
        "amount": None, "expenses": None, "revenue": None, "profit": None, "risk": None
    })


def _current_dataset_id(user_id: int) -> int | None:
    """Return the authenticated user's most recently processed upload only."""
    row = run_query(
        """SELECT id FROM uploaded_files
           WHERE user_id=%s AND status='processed'
           ORDER BY id DESC LIMIT 1""",
        (user_id,), fetchone=True,
    )["row"]
    return int(row["id"]) if row else None


def _dataset_rows_frame(user_id: int, dataset_id: int | None = None) -> pd.DataFrame:
    """Restore the cleaned row JSON for one user and optional upload."""
    rows = run_query(
        """SELECT row_data FROM dataset_rows
           WHERE user_id=%s AND (%s IS NULL OR uploaded_file_id=%s)
           ORDER BY uploaded_file_id, `row_number`""",
        (user_id, dataset_id, dataset_id), fetchall=True,
    )["rows"] or []
    records = []
    for row in rows:
        try:
            record = json.loads(row.get("row_data") or "")
        except (TypeError, ValueError):
            continue
        if isinstance(record, dict):
            records.append(record)
    return pd.DataFrame(records)


def _load_active_cleaned_dataset(user_id: int, dataset_id: int) -> pd.DataFrame:
    """Rebuild the active upload's cleaned frame and its models from owned rows."""
    try:
        restored = _dataset_rows_frame(user_id, dataset_id)
    except mysql.connector.Error:
        # A missing/unreachable durable row store must not be mistaken for a
        # generic upload with no canonical financial columns.
        raise
    except Exception as exc:
        print(f"[finsight] Could not restore dataset rows: {exc}")
        restored = pd.DataFrame()
    if not restored.empty:
        active = restored.copy()
        has_date = "tx_date" in active.columns
        if has_date:
            active["tx_date"] = pd.to_datetime(active["tx_date"], errors="coerce")
            training = active.dropna(subset=["tx_date"]).copy()
        else:
            training = pd.DataFrame()
        if has_date and not training.empty:
            training["Date_Number"] = (training["tx_date"] - training["tx_date"].min()).dt.days
            training["Month"] = training["tx_date"].dt.month
            training["Day_of_Week"] = training["tx_date"].dt.dayofweek
        session_data[user_id] = {
            # The full cleaned upload is the application data frame. Invalid
            # dates may prevent a forecast, but must not disappear from row
            # counts, analytics, or downloads.
            "df": active,
            "dataset_id": dataset_id,
            "models": {"amount": None, "expenses": None, "revenue": None, "profit": None, "risk": None},
        }
        if has_date and not training.empty:
            _train_models_for(user_id, training)
        _run_universal_analysis(user_id, active)
        return active

    rows = run_query(
        """SELECT tx_date, transaction_id, description, amount, revenue, expenses, profit,
                  customers, marketing_spend, tx_type, category, payment_method, department, city, status
           FROM financial_data
           WHERE user_id=%s AND uploaded_file_id=%s
           ORDER BY tx_date, id""",
        (user_id, dataset_id), fetchall=True,
    )["rows"] or []
    df = pd.DataFrame(rows)
    if df.empty:
        # Generic uploads have no canonical financial rows.  Record ownership
        # anyway so navigation does not repeatedly discard their restored
        # universal analysis and incorrectly show the upload-required state.
        session_data.setdefault(user_id, {})["dataset_id"] = dataset_id
        session_data.setdefault(user_id, {})["df"] = df
        return df
    df["tx_date"] = pd.to_datetime(df["tx_date"], errors="coerce")
    training = df.dropna(subset=["tx_date"]).copy()
    if not training.empty:
        training["Date_Number"] = (training["tx_date"] - training["tx_date"].min()).dt.days
        training["Month"] = training["tx_date"].dt.month
        training["Day_of_Week"] = training["tx_date"].dt.dayofweek
    session_data[user_id] = {
        "df": df,
        "dataset_id": dataset_id,
        "models": {"amount": None, "expenses": None, "revenue": None, "profit": None, "risk": None},
    }
    if not training.empty:
        _train_models_for(user_id, training)
    return df


def _active_dataset_and_frame(user_id: int) -> tuple[int | None, pd.DataFrame | None]:
    """Ensure the frame/model in memory belongs to this user's current upload."""
    dataset_id = _current_dataset_id(user_id)
    if dataset_id is None:
        return None, None
    state = session_data.get(user_id, {})
    if state.get("dataset_id") != dataset_id:
        return dataset_id, _load_active_cleaned_dataset(user_id, dataset_id)
    return dataset_id, state.get("df")


# =====================================================
# Universal analysis (dataset-agnostic predictions / insights)
# =====================================================
def _run_universal_analysis(user_id: int, raw_df: pd.DataFrame) -> dict | None:
    """Run universal analysis on the cleaned upload and cache it in memory."""
    if universal_analysis is None:
        return None
    try:
        clean = universal_analysis.clean_dataset(raw_df)
        analysis = universal_analysis.auto_analyze(clean)
        session_data.setdefault(user_id, {})["analysis"] = analysis
        session_data.setdefault(user_id, {})["analysis_df"] = clean.head(1000)
        return analysis
    except Exception as exc:
        print(f"[finsight] Universal analysis warning: {exc}")
        return None


def _analysis_context(user_id: int) -> dict | None:
    """Return the cached universal analysis (model objects removed)."""
    state = session_data.get(user_id, {})
    analysis = state.get("analysis")
    if isinstance(analysis, dict):
        try:
            sections = [
                universal_analysis.strip_model_keys(s)
                for s in analysis.get("sections", [])
            ]
        except Exception:
            sections = analysis.get("sections", [])
        return {
            "ok": analysis.get("ok", False),
            "rows": analysis.get("rows", 0),
            "columns": analysis.get("columns", 0),
            "types": analysis.get("types", {}),
            "sections": sections,
            "trends": analysis.get("trends", []),
            "insight": _localized_insight(analysis.get("insight", "")),
            "has_predictions": analysis.get("has_predictions", False),
            "message": analysis.get("message", ""),
        }
    return None


# =====================================================
# Groq chat assistant
# =====================================================
def _chat_number(value: object) -> str:
    try:
        number = float(value)
        if not np.isfinite(number):
            return "n/a"
        return f"{number:,.2f}"
    except (TypeError, ValueError):
        return "n/a"


def _chat_data_context(user_id: int) -> str:
    """Build a compact, user-scoped summary for the assistant.

    Only aggregate values, a small category ranking, and recent model output
    are sent to Groq. Raw uploads and credentials never leave this server.
    """
    try:
        dataset_id, frame = _active_dataset_and_frame(user_id)
    except mysql.connector.Error:
        raise
    except Exception as exc:
        app.logger.warning("Chat data context could not load the active dataset: %s", exc)
        return "No uploaded dataset is currently available."
    if dataset_id is None or frame is None or frame.empty:
        return "No uploaded dataset is currently available."

    working = frame.copy()
    lines = [f"Active dataset: {len(working):,} cleaned rows."]
    analysis = _analysis_context(user_id) or {}
    types = analysis.get("types") or _analysis_types_for(working)
    date_column = next((column for column, kind in types.items()
                        if kind == "date" and column in working.columns), None)
    if date_column:
        dates = pd.to_datetime(working[date_column], errors="coerce").dropna()
        if not dates.empty:
            lines.append(f"Date range: {dates.min().date()} to {dates.max().date()}.")
    numeric_columns = [column for column, kind in types.items()
                       if kind == "numeric" and column in working.columns]
    for metric in numeric_columns[:12]:
        values = pd.to_numeric(working[metric], errors="coerce").dropna()
        if not values.empty:
            lines.append(
                f"{metric.replace('_', ' ').title()}: total {_chat_number(values.sum())}; "
                f"average {_chat_number(values.mean())}."
            )

    value_column = numeric_columns[0] if numeric_columns else None
    category_column = next((column for column, kind in types.items()
                            if kind in {"categorical", "text", "boolean"}
                            and column in working.columns), None)
    if value_column and category_column:
        grouped = working[[category_column, value_column]].copy()
        grouped[value_column] = pd.to_numeric(grouped[value_column], errors="coerce")
        grouped[category_column] = grouped[category_column].fillna("Unspecified").astype(str).str.strip()
        grouped = grouped.dropna(subset=[value_column]).groupby(category_column)[value_column].sum()
        if not grouped.empty:
            top = grouped.sort_values(ascending=False).head(8)
            ranking = ", ".join(f"{name}: {_chat_number(value)}" for name, value in top.items())
            lines.append(f"Top {category_column.replace('_', ' ')} by {value_column}: {ranking}.")

    trends = analysis.get("trends") or []
    if trends:
        trend_text = "; ".join(
            f"{item.get('metric', 'metric')} {item.get('direction', 'stable')}"
            for item in trends[:6] if isinstance(item, dict)
        )
        if trend_text:
            lines.append(f"Detected trends: {trend_text}.")

    try:
        predictions = run_query(
            """SELECT prediction_type, prediction_date, predicted_value, model_name
               FROM predictions WHERE user_id=%s AND uploaded_file_id=%s
               ORDER BY prediction_id DESC LIMIT 8""",
            (user_id, dataset_id), fetchall=True,
        )["rows"] or []
        if predictions:
            forecast_text = "; ".join(
                f"{row.get('prediction_type')} on {row.get('prediction_date')}: "
                f"{_chat_number(row.get('predicted_value'))}"
                for row in predictions
            )
            lines.append(f"Recent forecast results: {forecast_text}.")
    except Exception as exc:
        app.logger.warning("Chat forecast context unavailable: %s", exc)

    try:
        risk = run_query(
            """SELECT risk_level, classification_date, explanation
               FROM risk_classifications WHERE user_id=%s AND uploaded_file_id=%s
               ORDER BY risk_id DESC LIMIT 1""",
            (user_id, dataset_id), fetchone=True,
        )["row"]
        if risk:
            explanation = str(risk.get("explanation") or "")[:300]
            lines.append(
                f"Latest risk classification: {risk.get('risk_level')} on "
                f"{risk.get('classification_date')}; {explanation}"
            )
    except Exception as exc:
        app.logger.warning("Chat risk context unavailable: %s", exc)

    return "\n".join(lines)[:10000]


def _openrouter_answer(messages: list[dict[str, str]]) -> str:
    """Call OpenRouter's OpenAI-compatible endpoint without exposing the key."""
    api_key = os.environ.get("OPENROUTER_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("The AI assistant is not configured on this deployment.")
    payload = json.dumps({
        "model": OPENROUTER_MODEL,
        "messages": messages,
        "temperature": 0.25,
        "max_tokens": 700,
    }).encode("utf-8")
    request_obj = Request(
        "https://openrouter.ai/api/v1/chat/completions",
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "HTTP-Referer": "https://finsightai-3ea6.onrender.com",
            "X-Title": "FinSight AI",
        },
        method="POST",
    )
    try:
        with urlopen(request_obj, timeout=GROQ_TIMEOUT) as response:
            body = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
            # The provider's message is useful in Render logs (invalid model,
            # quota, authentication, etc.), but never log request headers or
            # the API key. Keep the browser-facing response safe.
            detail = ""
            provider_code = ""
            try:
                error_body = json.loads(exc.read().decode("utf-8"))
                error_value = error_body.get("error") if isinstance(error_body, dict) else None
                if isinstance(error_value, dict):
                    detail = str(error_value.get("message") or "")
                    provider_code = str(error_value.get("code") or "")
                elif error_value:
                    detail = str(error_value)
            except (OSError, ValueError, TypeError):
                pass
            detail = re.sub(r"(?i)bearer\s+[A-Za-z0-9._-]+", "bearer [redacted]", detail)
            detail = re.sub(r"(?i)gsk_[A-Za-z0-9_-]+", "[redacted]", detail)
            app.logger.warning(
                "OpenRouter request returned HTTP %s for model %s (provider code %s): %s",
                exc.code, OPENROUTER_MODEL, provider_code or "unknown",
                detail[:500] or "no provider detail",
            )
            if exc.code == 401:
                raise RuntimeError("The AI assistant credentials are invalid. Please try again later.") from exc
            if exc.code in (401, 403):
                raise RuntimeError(
                    "The AI assistant key is invalid or the selected OpenRouter model is unavailable."
                ) from exc
            if exc.code == 429:
                raise RuntimeError("The AI assistant is temporarily busy. Please try again in a moment.") from exc
            raise RuntimeError("The AI assistant could not complete that request.") from exc
    except (URLError, TimeoutError) as exc:
        app.logger.warning("OpenRouter request failed: %s", type(exc).__name__)
        raise RuntimeError("The AI assistant is temporarily unavailable. Please try again.") from exc
    except (ValueError, OSError) as exc:
        app.logger.warning("OpenRouter response could not be read: %s", type(exc).__name__)
        raise RuntimeError("The AI assistant returned an invalid response.") from exc

    try:
        content = body["choices"][0]["message"]["content"]
    except (KeyError, IndexError, TypeError) as exc:
        app.logger.warning("Groq response did not contain assistant content")
        raise RuntimeError("The AI assistant returned an empty response.") from exc
    if not isinstance(content, str) or not content.strip():
        raise RuntimeError("The AI assistant returned an empty response.")
    return content.strip()[:12000]


# =====================================================
# Auth helpers / decorators
# =====================================================
def login_required(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if not _session_user_id():
            flash("Please log in to continue.", "warning")
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapper


@app.post("/api/chat")
@login_required
def chat():
    """Answer a natural-language question with the current user's context."""
    if not request.is_json:
        return jsonify({"error": "JSON body required."}), 400
    payload = request.get_json(silent=True) or {}
    message = str(payload.get("message") or "").strip()
    if not message:
        return jsonify({"error": "Enter a question for the AI assistant."}), 400
    if len(message) > 2000:
        return jsonify({"error": "Please keep your question under 2,000 characters."}), 400

    user_id = _session_user_id()
    locale = i18n.get_locale() if i18n is not None else "en"
    locale_names = getattr(i18n, "LOCALE_NAMES", {}) if i18n is not None else {}
    language = locale_names.get(locale, locale)
    system = (
        "You are FinSight AI, a concise and practical financial data-analysis assistant. "
        f"Answer in {language} ({locale}) because that is the user's selected language. "
        "Use the supplied dataset summary when present. Explain calculations plainly, "
        "state when information is unavailable, and do not invent figures, forecasts, "
        "or business facts. This is analytical guidance, not regulated financial advice. "
        "Keep answers helpful and readable with short paragraphs or bullets.\n\n"
        "DATA CONTEXT:\n" + _chat_data_context(user_id)
    )
    state = session_data.setdefault(user_id, {})
    history = state.setdefault("chat_messages", [])
    messages = [{"role": "system", "content": system}, *history[-10:],
                {"role": "user", "content": message}]
    try:
        answer = _openrouter_answer(messages)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 503
    history.extend([{"role": "user", "content": message},
                    {"role": "assistant", "content": answer}])
    state["chat_messages"] = history[-12:]
    return jsonify({"success": True, "message": answer, "locale": locale})


@app.post("/api/chat/reset")
@login_required
def reset_chat():
    """Forget only the current user's in-memory assistant conversation."""
    session_data.setdefault(_session_user_id(), {}).pop("chat_messages", None)
    return jsonify({"success": True})


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def _valid_email(value: str) -> bool:
    """Small, dependency-free validation before persisting login identifiers."""
    _, parsed = parseaddr(value)
    if parsed != value or len(value) > 254 or value.count("@") != 1:
        return False
    local, domain = value.rsplit("@", 1)
    # Keep this deliberately lightweight: reject malformed addresses without
    # imposing provider-specific rules on valid business email addresses.
    return bool(local and domain and "." in domain and not domain.startswith(".") and not domain.endswith("."))


def _password_matches(stored_hash: object, password: str) -> bool:
    """Verify a stored Werkzeug hash without allowing malformed data to raise."""
    if not isinstance(stored_hash, str) or not stored_hash or not password:
        return False
    try:
        return bool(check_password_hash(stored_hash, password))
    except (AttributeError, TypeError, ValueError):
        # A missing, truncated, or legacy non-Werkzeug value is not a valid
        # credential. Never turn it into a successful login or a server error.
        return False


def csrf_token() -> str:
    """Return a session-bound token for state-changing browser requests."""
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


@app.before_request
def protect_csrf():
    """Reject cross-site writes while allowing normal HTML forms and AJAX calls."""
    if request.method not in {"POST", "PUT", "PATCH", "DELETE"}:
        return None
    supplied = request.headers.get("X-CSRFToken") or request.form.get("csrf_token")
    expected = session.get("_csrf_token")
    if expected and supplied and secrets.compare_digest(expected, supplied):
        return None
    if request.is_json or request.path in {"/upload", "/upload/preview", "/risk-classify", "/powerbi/generate"}:
        return jsonify({"error": "Your session validation token is missing or expired. Refresh the page and try again."}), 400
    flash("Your form has expired. Please refresh the page and try again.", "warning")
    return redirect(request.referrer or url_for("index"))


@app.context_processor
def inject_security_context():
    return {"csrf_token": csrf_token, "format_number": _format_number}


# =====================================================
# Database initialization
# =====================================================
def init_database() -> None:
    """Create or upgrade the complete schema without modifying existing data."""
    try:
        create_database_tables()
        print("[finsight] Database schema ready.")
    except mysql.connector.Error as err:
        app.logger.error(
            "Database schema initialization failed: %s: %s",
            type(err).__name__,
            err,
            exc_info=(type(err), err, err.__traceback__),
        )
    except (RuntimeError, ValueError) as err:
        app.logger.error("Database schema configuration failed: %s", err)


# =====================================================
# Column normalisation for uploaded files
# =====================================================
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Backward-compatible wrapper around the conservative business mapper."""
    return apply_mapping(df, detect_columns(df).get("mapping", {}))


def _excel_first_row_is_data(frame: pd.DataFrame) -> bool:
    """Detect an Excel sheet whose first row is data rather than a header."""
    if frame is None or frame.empty:
        return False

    header_values = pd.Series(list(frame.columns), dtype="string").str.strip()
    header_numeric = pd.to_numeric(header_values, errors="coerce").notna()
    header_blank = header_values.eq("") | header_values.str.match(r"^Unnamed", case=False, na=False)
    if not bool((header_numeric | header_blank).any()):
        # A headerless sheet can also start with text-only values. If several
        # apparent headers repeat in the following rows, they are data values
        # rather than field names. Require a small signal so a normal header
        # such as Status/Category is not reclassified by accident.
        aliases = {
            alias.casefold()
            for spec in FIELD_SPECS.values()
            for alias in spec.get("aliases", [])
        }
        header_names = {
            re.sub(r"[^a-z0-9]+", "_", str(column).strip().casefold()).strip("_")
            for column in frame.columns
        }
        alias_hits = len(header_names & aliases)
        required_signal = max(2, (len(header_names) + 4) // 5)
        if alias_hits >= required_signal:
            return False
        repeat_hits = 0
        for column in frame.columns:
            header = re.sub(r"[^a-z0-9]+", "_", str(column).strip().casefold()).strip("_")
            if not header:
                continue
            values = {
                re.sub(r"[^a-z0-9]+", "_", str(value).strip().casefold()).strip("_")
                for value in frame[column].head(20).dropna()
            }
            if header in values:
                repeat_hits += 1
        if repeat_hits >= required_signal:
            return True
        return False

    first_row = frame.iloc[0]
    non_empty = first_row.dropna()
    if non_empty.empty:
        return True
    numeric_count = int(pd.to_numeric(non_empty, errors="coerce").notna().sum())
    # A normal header row is textual. Multiple numeric values in the first row
    # are therefore a strong signal that pandas consumed a real data row as the
    # header (the common shape of an uncleaned Excel export).
    return numeric_count >= 2


def _read_excel_upload(source: str | Path | io.BytesIO) -> pd.DataFrame:
    """Read normal Excel headers while preserving headerless data exports."""
    frame = pd.read_excel(source)
    if not _excel_first_row_is_data(frame):
        return frame
    if hasattr(source, "seek"):
        source.seek(0)
    frame = pd.read_excel(source, header=None)
    frame.columns = [f"column_{index + 1}" for index in range(len(frame.columns))]
    # Keep this small piece of provenance for the preview and cleaning report.
    # Generic names are safer than guessing that an arbitrary numeric column is
    # revenue, date, or another business field.
    frame.attrs["finsight_headerless"] = True
    return frame


def _read_upload_dataframe(source: str | Path | io.BytesIO, filename: str) -> pd.DataFrame:
    """Read one supported upload format into a bounded dataframe."""
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        try:
            frame = pd.read_csv(source)
            if len(frame.columns) == 1:
                # A semicolon-delimited export is common in European tools.
                if hasattr(source, "seek"):
                    source.seek(0)
                frame = pd.read_csv(source, sep=None, engine="python")
        except UnicodeDecodeError:
            if hasattr(source, "seek"):
                source.seek(0)
            frame = pd.read_csv(source, encoding="latin-1", sep=None, engine="python")
    elif extension in {".xlsx", ".xls"}:
        frame = _read_excel_upload(source)
    elif extension == ".json":
        if hasattr(source, "seek"):
            source.seek(0)
        if isinstance(source, (str, Path)):
            with open(source, "r", encoding="utf-8") as json_file:
                payload = json.load(json_file)
        else:
            payload = json.load(source)
        frame = profile_json_payload(payload)
    else:
        raise ValueError("File must be CSV, XLSX, XLS, or JSON.")

    if frame.empty or len(frame.columns) == 0:
        raise ValueError("File is empty or has no tabular data.")
    if len(frame.columns) > MAX_DATA_COLUMNS:
        raise ValueError(f"The file contains too many columns. Maximum allowed is {MAX_DATA_COLUMNS}.")
    if len(frame) > 1_000_000:
        raise ValueError("The file contains too many rows. Maximum allowed is 1,000,000.")
    return frame


def _mapping_from_request() -> dict[str, str]:
    raw = request.form.get("mapping")
    if not raw:
        return {}
    try:
        payload = json.loads(raw)
    except (TypeError, ValueError) as exc:
        raise ValueError("The column mapping is invalid.") from exc
    if not isinstance(payload, dict):
        raise ValueError("The column mapping is invalid.")
    return {str(key): str(value).strip() for key, value in payload.items() if value}


# These opt-in examples are only returned by the explicit template-download
# route. They are never used as a fallback source for an uploaded dataset.
BUSINESS_TEMPLATE_ROWS = {
    "general": [
        {"Date": "2026-01-05", "Transaction_ID": "GEN-1001", "Description": "Product sale", "Revenue": 12500, "Amount": 12500, "Expenses": 7300, "Profit": 5200, "Customers": 84, "Category": "Sales", "Department": "Commercial", "Payment_Method": "Card", "City": "Berlin", "Status": "Completed"},
        {"Date": "2026-02-05", "Transaction_ID": "GEN-1002", "Description": "Subscription renewal", "Revenue": 14800, "Amount": 14800, "Expenses": 8100, "Profit": 6700, "Customers": 96, "Category": "Subscriptions", "Department": "Commercial", "Payment_Method": "Transfer", "City": "Berlin", "Status": "Completed"},
        {"Date": "2026-03-05", "Transaction_ID": "GEN-1003", "Description": "Product sale", "Revenue": 16100, "Amount": 16100, "Expenses": 8950, "Profit": 7150, "Customers": 103, "Category": "Sales", "Department": "Commercial", "Payment_Method": "Card", "City": "Munich", "Status": "Completed"},
    ],
    "retail": [
        {"Date": "2026-01-10", "Transaction_ID": "RTL-2001", "Description": "Store order", "Revenue": 4200, "Amount": 4200, "Expenses": 2550, "Profit": 1650, "Customers": 32, "Category": "Electronics", "Department": "Store", "Payment_Method": "Card", "City": "Berlin", "Status": "Completed"},
        {"Date": "2026-02-10", "Transaction_ID": "RTL-2002", "Description": "Online order", "Revenue": 5100, "Amount": 5100, "Expenses": 2990, "Profit": 2110, "Customers": 39, "Category": "Home", "Department": "E-commerce", "Payment_Method": "PayPal", "City": "Hamburg", "Status": "Completed"},
        {"Date": "2026-03-10", "Transaction_ID": "RTL-2003", "Description": "Store order", "Revenue": 5750, "Amount": 5750, "Expenses": 3380, "Profit": 2370, "Customers": 44, "Category": "Electronics", "Department": "Store", "Payment_Method": "Card", "City": "Berlin", "Status": "Completed"},
    ],
    "service": [
        {"Date": "2026-01-15", "Transaction_ID": "SRV-3001", "Description": "Consulting engagement", "Revenue": 8600, "Amount": 8600, "Expenses": 4100, "Profit": 4500, "Customers": 7, "Category": "Consulting", "Department": "Delivery", "Payment_Method": "Transfer", "City": "Frankfurt", "Status": "Completed"},
        {"Date": "2026-02-15", "Transaction_ID": "SRV-3002", "Description": "Support retainer", "Revenue": 9200, "Amount": 9200, "Expenses": 4350, "Profit": 4850, "Customers": 9, "Category": "Support", "Department": "Customer Success", "Payment_Method": "Transfer", "City": "Frankfurt", "Status": "Completed"},
        {"Date": "2026-03-15", "Transaction_ID": "SRV-3003", "Description": "Implementation project", "Revenue": 11200, "Amount": 11200, "Expenses": 5660, "Profit": 5540, "Customers": 6, "Category": "Implementation", "Department": "Delivery", "Payment_Method": "Card", "City": "Cologne", "Status": "Completed"},
    ],
}


# =====================================================
# Per-user Power BI Desktop resources
# =====================================================
def _powerbi_resource(user_id: int, create: bool = True) -> dict | None:
    """Return the authenticated user's opaque Power BI resource record."""
    result = run_query("SELECT * FROM user_powerbi_resources WHERE user_id=%s", (user_id,), fetchone=True)
    row = result["row"]
    if row or not create:
        return row
    folder_token = secrets.token_urlsafe(18).replace("-", "").replace("_", "")
    run_query(
        "INSERT INTO user_powerbi_resources (user_id, folder_token) VALUES (%s, %s)",
        (user_id, folder_token), commit=True,
    )
    return run_query("SELECT * FROM user_powerbi_resources WHERE user_id=%s", (user_id,), fetchone=True)["row"]


def _powerbi_paths(resource: dict, user_id: int, dataset_id: int | None = None) -> dict[str, Path]:
    root = POWERBI_ROOT / resource["folder_token"] / "powerbi"
    if dataset_id is not None:
        root = root / f"dataset_{dataset_id}"
    data = root / "data"
    pbix_name = f"finsight_user_{user_id}_dataset_{dataset_id or 'latest'}.pbix"
    return {"root": root, "data": data, "financial": data / "financial_data.csv",
            "predictions": data / "predictions.csv", "pbix": root / pbix_name,
            "readme": root / "README-PowerBI-Desktop.txt"}


def _next_dataset_version(user_id: int) -> int:
    row = run_query(
        "SELECT COALESCE(MAX(version), 0) + 1 AS next_version "
        "FROM uploaded_files WHERE user_id=%s", (user_id,), fetchone=True,
    )["row"]
    return int(row["next_version"] if row else 1)


# =====================================================
# History helpers
# =====================================================
def add_history(user_id: int, company_id: int, event_type: str, title: str,
                *, file_id: int | None = None, prediction_id: int | None = None,
                status: str = "ok", details: str | None = None) -> None:
    run_query(
        """
        INSERT INTO dashboard_history
            (user_id, company_id, event_type, event_title, file_id, prediction_id, status, details)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (user_id, company_id, event_type, title, file_id, prediction_id, status, details),
        commit=True,
    )


# =====================================================
# Routes - Auth
# =====================================================
@app.route("/")
def index():
    if _session_user_id():
        return redirect(url_for("dashboard"))
    return render_template("index.html")


@app.get("/healthz")
def healthz():
    """Lightweight deployment health check that does not require a DB round trip."""
    return jsonify({"status": "ok", "service": "finsight-ai"})


@app.get("/templates/<template_name>/<file_format>")
@login_required
def download_business_template(template_name: str, file_format: str):
    """Download a small, valid business template accepted by the upload flow."""
    if template_name not in BUSINESS_TEMPLATE_ROWS or file_format not in {"csv", "xlsx", "json"}:
        return jsonify({"error": "Template not found."}), 404
    frame = pd.DataFrame(BUSINESS_TEMPLATE_ROWS[template_name])
    safe_name = f"finsight_{template_name}_template.{file_format}"
    if file_format == "csv":
        output = io.BytesIO(frame.to_csv(index=False).encode("utf-8"))
        mimetype = "text/csv"
    elif file_format == "json":
        output = io.BytesIO(frame.to_json(orient="records", date_format="iso").encode("utf-8"))
        mimetype = "application/json"
    else:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            frame.to_excel(writer, index=False, sheet_name="Business_Data")
        output.seek(0)
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return send_file(output, as_attachment=True, download_name=safe_name, mimetype=mimetype)


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        company_name = (request.form.get("company_name") or "").strip()

        if not name or not email or not password or not company_name:
            flash("All fields are required.", "danger")
            return render_template("signup.html")

        if not _valid_email(email):
            flash("Enter a valid email address.", "danger")
            return render_template("signup.html")

        if len(name) > 120 or len(company_name) > 255:
            flash("Name or company name is too long.", "danger")
            return render_template("signup.html")

        if len(password) < 6:
            flash("Password must be at least 6 characters long.", "danger")
            return render_template("signup.html")

        try:
            user_columns = _load_user_columns()
            email_column = user_columns.get("email")
            password_column = user_columns.get("password")
        except (mysql.connector.Error, RuntimeError) as exc:
            app.logger.error("Signup schema lookup failed: %s", exc, exc_info=True)
            flash("Database connection failed. Please try again later.", "danger")
            return render_template("signup.html"), 503
        if not email_column or not password_column:
            flash("The users table is missing its email or password field.", "danger")
            return render_template("signup.html"), 503

        # Check whether the email is already used. Keep connection failures in
        # the signup flow so users receive a useful message instead of a 500.
        try:
            existing = run_query(
                f"SELECT id FROM users WHERE {_quote_identifier(email_column)} = %s",
                (email,), fetchone=True,
            )
        except Exception as exc:
            app.logger.error("Signup email lookup failed: %s", exc, exc_info=True)
            flash("The database is currently unavailable. Please try again in a moment.", "danger")
            return render_template("signup.html"), 503
        if existing and existing["row"]:
            flash("This email is already registered. Please log in.", "warning")
            return redirect(url_for("login"))

        conn = None
        cur = None
        try:
            conn = get_db()
            cur = conn.cursor(dictionary=True)
            # Find or create the company
            cur.execute("SELECT id FROM companies WHERE company_name = %s", (company_name,))
            row = cur.fetchone()
            if row:
                company_id = row["id"]
            else:
                cur.execute(
                    "INSERT INTO companies (company_name) VALUES (%s)",
                    (company_name,),
                )
                company_id = cur.lastrowid

            name_parts = name.split(maxsplit=1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ""
            insert_values = []
            insert_columns = []
            display_column = user_columns.get("display")
            if display_column:
                insert_columns.append(display_column)
                insert_values.append(name)
            # Some existing deployments contain both a legacy full-name
            # column and required firstName/lastName columns. Populate every
            # supported name field so the live schema's required fields are
            # satisfied instead of relying on a default value.
            if user_columns.get("first_name") and user_columns["first_name"] not in insert_columns:
                insert_columns.append(user_columns["first_name"])
                insert_values.append(first_name)
            if user_columns.get("last_name") and user_columns["last_name"] not in insert_columns:
                insert_columns.append(user_columns["last_name"])
                insert_values.append(last_name)
            insert_columns.extend([email_column, password_column])
            insert_values.extend([email, generate_password_hash(password)])
            if user_columns.get("company_id"):
                insert_columns.append(user_columns["company_id"])
                insert_values.append(company_id)
            if user_columns.get("role"):
                insert_columns.append(user_columns["role"])
                insert_values.append("user")
            if not display_column and not user_columns.get("first_name") and not user_columns.get("last_name"):
                raise RuntimeError("The users table has no supported name field.")
            quoted_columns = ", ".join(_quote_identifier(column) for column in insert_columns)
            placeholders = ", ".join(["%s"] * len(insert_values))
            cur.execute(
                f"INSERT INTO users ({quoted_columns}) VALUES ({placeholders})",
                tuple(insert_values),
            )
            user_id = cur.lastrowid
            conn.commit()
        except Exception as exc:
            if conn is not None:
                try:
                    conn.rollback()
                except Exception:
                    pass
            app.logger.error("Signup failed: %s", exc, exc_info=True)
            if isinstance(exc, mysql.connector.Error) and getattr(exc, "errno", None) == errorcode.ER_ACCESS_DENIED_ERROR:
                flash("Cannot connect to MySQL: access denied. "
                      "Check DB_USER and DB_PASSWORD in your environment configuration.", "danger")
            elif isinstance(exc, mysql.connector.Error) and getattr(exc, "errno", None) == errorcode.ER_DUP_ENTRY:
                flash("This email is already registered. Please log in.", "warning")
            elif isinstance(exc, RuntimeError):
                flash("The database schema is missing a required account field. Please contact the administrator.", "danger")
            elif isinstance(exc, mysql.connector.Error):
                flash("The database is currently unavailable or could not save your account. Please try again in a moment.", "danger")
            else:
                flash("We could not create your account right now. Please try again.", "danger")
            return render_template("signup.html")
        finally:
            if cur is not None:
                cur.close()
            if conn is not None:
                conn.close()

        # Account creation is already committed at this point. These are
        # optional conveniences and must not turn a successful signup into a
        # false failure when a legacy database is missing one of their tables.
        try:
            _powerbi_resource(user_id)
        except Exception as exc:
            app.logger.warning("Signup Power BI resource setup skipped: %s", exc)
        try:
            add_history(user_id, company_id, "signup", f"New account created for {company_name}")
        except Exception as exc:
            app.logger.warning("Signup history entry skipped: %s", exc)

        session["user_id"] = user_id
        session["user_name"] = name
        session["company_id"] = company_id
        session["company_name"] = company_name
        flash(f"Welcome to FinSight AI, {name}!", "success")
        return redirect(url_for("dashboard"))

    return render_template("signup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""

        if not email or not password:
            flash("Email and password are required.", "danger")
            return render_template("login.html")

        if not _valid_email(email):
            flash("Enter a valid email address.", "danger")
            return render_template("login.html")

        try:
            user_columns = _load_user_columns()
            email_column = user_columns.get("email")
            password_column = user_columns.get("password")
            if not email_column or not password_column:
                flash("The users table is missing its email or password field.", "danger")
                return render_template("login.html")

            company_column = user_columns.get("company_id")
            if company_column:
                company_sql = f"u.{_quote_identifier(company_column)} AS company_id, c.company_name"
                # A user row must not disappear merely because its company
                # lookup is unavailable; the checked-in schema enforces the
                # relationship, while LEFT JOIN keeps auth failure explicit
                # for any pre-existing data that does not.
                company_join = "LEFT JOIN companies c ON c.id = u." + _quote_identifier(company_column)
            else:
                company_sql = "NULL AS company_id, '' AS company_name"
                company_join = ""
            query_result = run_query(
                f"""
                SELECT u.id, {_user_display_expression(user_columns)},
                       u.{_quote_identifier(password_column)} AS password_hash,
                       {company_sql}
                FROM users u
                {company_join}
                WHERE u.{_quote_identifier(email_column)} = %s
                """,
                (email,),
                fetchone=True,
            )
            row = query_result.get("row") if isinstance(query_result, dict) else None
        except mysql.connector.Error as exc:
            app.logger.error(
                "Login database lookup failed: %s: %s",
                type(exc).__name__,
                exc,
                exc_info=(type(exc), exc, exc.__traceback__),
            )
            flash("Database connection failed. Please try again later.", "danger")
            return render_template("login.html"), 503
        except RuntimeError as exc:
            app.logger.error("Login database configuration failed: %s", exc)
            flash("Database connection failed. Please try again later.", "danger")
            return render_template("login.html"), 503

        if not isinstance(row, dict) or not row.get("id") or not _password_matches(row.get("password_hash"), password):
            flash("Invalid email or password.", "danger")
            return render_template("login.html")

        display_name = row.get("name") or email
        session.clear()
        session["user_id"] = row["id"]
        session["user_name"] = display_name
        session["company_id"] = row.get("company_id")
        session["company_name"] = row.get("company_name") or ""
        flash(f"Welcome back, {display_name}!", "success")
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.post("/logout")
def logout():
    user_id = _session_user_id()
    if user_id is not None:
        session_data.pop(user_id, None)
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


# =====================================================
# Routes - Dashboard
# =====================================================
@app.route("/dashboard")
@login_required
def dashboard():
    user_id = _session_user_id()
    company_id = _session_company_id()

    # Pull recent uploads
    uploads = run_query(
        """
        SELECT id, original_name, rows_imported, status, created_at
        FROM uploaded_files
        WHERE user_id = %s
        ORDER BY created_at DESC
        LIMIT 5
        """,
        (user_id,),
        fetchall=True,
    )["rows"] or []

    active_dataset_id, df = _active_dataset_and_frame(user_id)

    # Recent predictions for the active upload only. A previous upload must
    # not make the current dataset look as if it has forecasts.
    predictions = run_query(
        """
        SELECT prediction_id, prediction_type, prediction_date, predicted_value, created_at
        FROM predictions
        WHERE user_id = %s AND uploaded_file_id = %s
        ORDER BY created_at DESC
        LIMIT 5
        """,
        (user_id, active_dataset_id),
        fetchall=True,
    )["rows"] or []
    risk_status = None
    if active_dataset_id is not None:
        risk_status = run_query(
            """SELECT risk_level, classification_date, explanation FROM risk_classifications
               WHERE user_id=%s AND uploaded_file_id=%s ORDER BY risk_id DESC LIMIT 1""",
            (user_id, active_dataset_id), fetchone=True,
        )["row"]

    analysis_context = _analysis_context(user_id) or {}
    types = analysis_context.get("types") or _analysis_types_for(df)
    numeric_metrics = []
    if df is not None and not df.empty:
        for column, kind in types.items():
            if kind != "numeric" or column not in df.columns:
                continue
            values = pd.to_numeric(df[column], errors="coerce")
            if values.notna().any():
                if _visual_identifier_like(str(column), values):
                    continue
                additive = _visual_additive_measure(str(column))
                numeric_metrics.append({
                    "name": column,
                    "label": f"{str(column).replace('_', ' ').title()} {'total' if additive else 'average'}",
                    "value": float(values.sum() if additive else values.mean()),
                })
    date_range = ""
    if df is not None and not df.empty:
        date_columns = [column for column, kind in types.items() if kind == "date" and column in df.columns]
        if date_columns:
            dates = pd.to_datetime(df[date_columns[0]], errors="coerce").dropna()
            if not dates.empty:
                date_range = f"{dates.min().date()} to {dates.max().date()}"
    stats = {
        "rows": int(len(df)) if df is not None else 0,
        "columns": int(len(df.columns)) if df is not None else 0,
        "numeric_metrics": numeric_metrics[:3],
        "date_range": date_range,
        "has_data": bool(df is not None and not df.empty),
    }
    return render_template(
        "dashboard.html",
        company_name=session["company_name"],
        user_name=session["user_name"],
        uploads=uploads,
        predictions=predictions,
        stats=stats,
        risk_status=risk_status,
        analysis=analysis_context,
        insight=analysis_context.get("insight", ""),
        analysis_trends=analysis_context.get("trends", []),
    )


def _visual_text(value: object) -> str:
    """Render an uploaded dimension value without replacing it with a fake label."""
    if value is None:
        return "Unspecified"
    try:
        missing = pd.isna(value)
        if isinstance(missing, (bool, np.bool_)) and missing:
            return "Unspecified"
    except (TypeError, ValueError):
        pass
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value).strip() or "Unspecified"


def _visual_identifier_like(column: str, values: pd.Series) -> bool:
    """Keep IDs and row counters out of measure charts when their values are numeric."""
    normalized = re.sub(r"[^a-z0-9]+", "_", str(column).casefold()).strip("_")
    identifier_name = (
        normalized in {"id", "index", "row", "row_number", "number", "code", "key"}
        or normalized.endswith(("_id", "_index", "_number", "_code", "_key"))
    )
    numeric = pd.to_numeric(values, errors="coerce").dropna()
    if numeric.empty:
        return False
    unique_ratio = float(numeric.nunique() / len(numeric))
    if identifier_name and unique_ratio >= 0.85:
        return True
    # Headerless spreadsheets often expose a generated ``column_N`` name. A
    # nearly unique monotonic counter is still an identifier even without a
    # semantic header; it should not become the default business measure.
    if normalized.startswith("column_") and unique_ratio >= 0.9 and numeric.is_monotonic_increasing:
        return True
    return False


def _visual_additive_measure(column: str) -> bool:
    """Return whether summing a measure is a useful dashboard total."""
    normalized = re.sub(r"[^a-z0-9]+", "_", str(column).casefold()).strip("_")
    tokens = (
        "revenue", "sales", "income", "amount", "expense", "expenses", "cost",
        "profit", "price", "value", "quantity", "qty", "count", "orders",
        "units", "volume", "spend", "budget", "balance",
    )
    return any(token in normalized for token in tokens)


def _visualization_data(df: pd.DataFrame, types: dict[str, str] | None = None) -> dict[str, object]:
    """Select real dimensions/measures and aggregate only uploaded values."""
    empty = {
        "numeric_columns": [], "date_column": None, "metric_column": None,
        "dimension_column": None, "category_label": "Category",
        "category_metric_label": "Value", "categories": [], "column_data": [],
        "area_data": [], "locations": [], "totals": {},
        "area_message": "No valid date/time column available for an area chart.",
        "donut_message": "No suitable categorical data available for a donut chart.",
        "column_message": "No suitable categorical data available for a column chart.",
        "distribution_data": [], "distribution_label": "Value",
        "distribution_message": "No numeric distribution is available.",
        "map_message": "No geographic data available for map visualization.",
    }
    if df is None or df.empty:
        return empty
    work = df.copy()
    types = types or _analysis_types_for(work, use_cached=False)
    numeric_columns: list[str] = []
    for column in work.columns:
        if types.get(column) != "numeric":
            continue
        values = pd.to_numeric(work[column], errors="coerce")
        if values.notna().any() and not _visual_identifier_like(str(column), values):
            work[column] = values
            numeric_columns.append(str(column))
    empty["numeric_columns"] = numeric_columns
    if not numeric_columns:
        empty["area_message"] = "No suitable numeric data available for an area chart."
        empty["map_message"] = "A map needs a numeric measure in addition to geographic values."
        # Categorical-only uploads still have a useful visualization: count
        # rows by the best available dimension instead of inventing a revenue
        # or monetary measure.
        dimensions = []
        for column in work.columns:
            kind = types.get(column)
            if kind not in {"categorical", "text", "boolean"}:
                continue
            values = work[column].dropna()
            try:
                unique_count = int(values.astype("string").nunique())
            except Exception:
                unique_count = 0
            if 2 <= unique_count <= 12:
                normalized = str(column).casefold().replace(" ", "_")
                rank = next((index for index, token in enumerate(
                    ("category", "segment", "department", "product", "region", "country", "state", "city", "type")
                ) if token in normalized), 99)
                dimensions.append((rank, -unique_count, str(column)))
        dimensions.sort()
        dimension = dimensions[0][2] if dimensions else None
        if dimension:
            counted = work[dimension].map(_visual_text).value_counts().head(12)
            categories = [
                {"category": str(category), "value": int(value)}
                for category, value in counted.items()
            ]
            empty["dimension_column"] = dimension
            empty["category_label"] = str(dimension).replace("_", " ").title()
            empty["category_metric_label"] = "Rows"
            empty["column_data"] = categories
            empty["column_message"] = "" if categories else "No suitable categorical data available for a column chart."
            donut_categories = categories if 2 <= len(categories) <= 6 else []
            empty["categories"] = donut_categories
            empty["donut_message"] = "" if donut_categories else "A donut chart needs 2–6 categorical parts."
        return empty

    metric_priority = ("sales", "revenue", "income", "amount", "profit", "expense",
                       "cost", "customers", "quantity", "count", "value")
    def metric_rank(column: str) -> tuple[int, int, float, int, str]:
        normalized = str(column).casefold().replace(" ", "_")
        rank = next((index for index, token in enumerate(metric_priority) if token in normalized), 999)
        values = pd.to_numeric(work[column], errors="coerce").dropna()
        variation = float(values.std(ddof=0) / max(abs(float(values.mean())), 1e-9)) if len(values) > 1 else 0.0
        # Prefer a field with enough distinct observations to chart a real
        # measure; low-cardinality month/day/status codes are dimensions.
        return rank, -int(values.nunique()), -variation, -int(values.notna().sum()), str(column)
    numeric_columns.sort(key=metric_rank)
    metric = numeric_columns[0]
    empty["metric_column"] = metric

    dimensions = []
    for column in work.columns:
        kind = types.get(column)
        if kind not in {"categorical", "text", "boolean"}:
            continue
        values = work[column].dropna()
        try:
            unique_count = int(values.astype("string").nunique())
        except Exception:
            unique_count = 0
        if unique_count < 2 or unique_count > 12:
            continue
        normalized = str(column).casefold().replace(" ", "_")
        rank = next((index for index, token in enumerate(
            ("category", "segment", "department", "product", "region", "country", "state", "city", "type")
        ) if token in normalized), 99)
        dimensions.append((rank, -unique_count, str(column)))
    dimensions.sort()
    dimension = dimensions[0][2] if dimensions else None
    empty["dimension_column"] = dimension
    if dimension:
        grouped = work[[dimension, metric]].copy()
        grouped[dimension] = grouped[dimension].map(_visual_text)
        grouped[metric] = pd.to_numeric(grouped[metric], errors="coerce")
        grouped = grouped.dropna(subset=[metric])
        grouped = grouped.groupby(dimension, as_index=False)[metric].sum(min_count=1)
        grouped = grouped.rename(columns={dimension: "category", metric: "value"})
        grouped = grouped.replace([np.inf, -np.inf], np.nan).dropna(subset=["value"])
        grouped = grouped.sort_values("value", ascending=False).head(12)
        categories = grouped.to_dict("records")
        empty["categories"] = categories
        empty["column_data"] = categories
        empty["category_label"] = str(dimension).replace("_", " ").title()
        empty["category_metric_label"] = f"Total {str(metric).replace('_', ' ').title()}"
        donut_categories = (
            categories if len(categories) <= 6
            and all(float(row["value"]) >= 0 for row in categories)
            and sum(float(row["value"]) for row in categories) > 0 else []
        )
        empty["categories"] = donut_categories
        empty["donut_message"] = "" if donut_categories else "A donut chart needs 2–6 non-negative parts."
        empty["column_message"] = "" if categories else "No suitable categorical data available for a column chart."

    date_column = _date_column_for(work, types)
    empty["date_column"] = date_column
    if date_column:
        dated = work[[date_column, metric]].copy()
        dated["_date"] = pd.to_datetime(dated[date_column], errors="coerce")
        dated[metric] = pd.to_numeric(dated[metric], errors="coerce")
        dated = dated.dropna(subset=["_date", metric]).sort_values("_date")
        if dated["_date"].nunique() >= 2:
            if dated["_date"].nunique() > 120:
                dated["period"] = dated["_date"].dt.to_period("M").astype(str)
                area = dated.groupby("period", as_index=False)[metric].sum(min_count=1)
            else:
                dated["period"] = dated["_date"].dt.strftime("%Y-%m-%d")
                area = dated.groupby("period", as_index=False)[metric].sum(min_count=1)
            area = area.rename(columns={metric: "value"})
            empty["area_data"] = area.to_dict("records")
            empty["area_message"] = ""

    geo_priority = ("country_name", "country", "nation", "state", "province", "region", "city", "location")
    geo_candidates = []
    for column in work.columns:
        kind = types.get(column)
        if kind not in {"categorical", "text"}:
            continue
        values = work[column].dropna()
        unique_count = int(values.astype("string").nunique()) if not values.empty else 0
        if not 0 < unique_count <= 200:
            continue
        normalized = str(column).casefold().replace(" ", "_")
        rank = next((index for index, token in enumerate(geo_priority)
                     if normalized == token or normalized.endswith("_" + token)), None)
        if rank is not None:
            geo_candidates.append((rank, -unique_count, str(column)))
    geo_candidates.sort()
    geo_column = geo_candidates[0][2] if geo_candidates else None
    if geo_column:
        locations = work[[geo_column, metric]].copy()
        locations["location"] = locations[geo_column].map(_visual_text)
        locations["value"] = pd.to_numeric(locations[metric], errors="coerce")
        locations = locations.dropna(subset=["value"])
        locations = locations.groupby("location", as_index=False)["value"].sum(min_count=1)
        locations = locations.sort_values("value", ascending=False).head(50)
        empty["locations"] = locations.to_dict("records")
        empty["map_message"] = "" if not locations.empty else "No geographic data available for map visualization."

    additive_columns = [column for column in numeric_columns if _visual_additive_measure(column)]
    empty["totals"] = {
        column: float(pd.to_numeric(work[column], errors="coerce").sum())
        for column in additive_columns
    }

    # A histogram is more honest for generic measures such as age, score, or
    # temperature than a bar chart of unrelated totals. It is shown only when
    # the selected metric has enough variation to form meaningful bins.
    metric_values = pd.to_numeric(work[metric], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    if metric_values.nunique() >= 5:
        try:
            bin_count = max(4, min(8, int(np.ceil(np.sqrt(len(metric_values))))))
            counts, edges = np.histogram(metric_values.to_numpy(dtype=float), bins=bin_count)
            distribution = []
            for index, count in enumerate(counts):
                distribution.append({
                    "range": f"{edges[index]:g}–{edges[index + 1]:g}",
                    "value": int(count),
                })
            empty["distribution_data"] = distribution
            empty["distribution_label"] = str(metric).replace("_", " ").title()
            empty["distribution_message"] = ""
        except (TypeError, ValueError):
            pass
    return empty


@app.route("/analytics")
@app.route("/visualizations")
@login_required
def analytics():
    """Render real visualizations derived from the active uploaded dataset."""
    user_id = _session_user_id()
    _, df = _active_dataset_and_frame(user_id)
    empty_context = {
        "company_name": session["company_name"], "has_data": False,
        "monthly": [], "area_data": [], "categories": [], "column_data": [],
        "locations": [], "cities": [], "totals": {}, "distribution_data": [],
        "distribution_label": "Value", "trend_keys": [],
        "trend_labels": [], "category_label": "Category", "has_trend": False,
        "has_area": False, "has_categories": False, "has_column": False,
        "has_totals": False, "has_distribution": False, "has_map": False, "has_cities": False,
        "category_metric_label": "Value", "generic_analytics": True,
        "area_message": "No valid date/time column available for an area chart.",
        "donut_message": "No suitable categorical data available for a donut chart.",
        "column_message": "No suitable categorical data available for a column chart.",
        "distribution_message": "No numeric distribution is available.",
        "map_message": "No geographic data available for map visualization.",
    }
    if df is None or df.empty:
        return render_template("analytics.html", **empty_context)
    visual = _visualization_data(df)
    return render_template(
        "analytics.html", company_name=session["company_name"], has_data=True,
        monthly=visual["area_data"], area_data=visual["area_data"],
        categories=visual["categories"], column_data=visual["column_data"],
        locations=visual["locations"], cities=visual["locations"], totals=visual["totals"],
        distribution_data=visual["distribution_data"], distribution_label=visual["distribution_label"],
        generic_analytics=True, trend_keys=["value"] if visual["area_data"] else [],
        trend_labels=[str(visual["metric_column"]).replace("_", " ").title()] if visual["metric_column"] else [],
        category_label=visual["category_label"], category_metric_label=visual["category_metric_label"],
        has_trend=bool(visual["area_data"]), has_area=bool(visual["area_data"]),
        has_categories=bool(visual["categories"]), has_column=bool(visual["column_data"]),
        has_totals=bool(visual["totals"]), has_distribution=bool(visual["distribution_data"]),
        has_map=bool(visual["locations"]),
        has_cities=bool(visual["locations"]), area_message=visual["area_message"],
        donut_message=visual["donut_message"], column_message=visual["column_message"],
        distribution_message=visual["distribution_message"],
        map_message=visual["map_message"],
    )


_GEOCODE_CACHE: dict[str, tuple[float, float] | None] = {}


def _geocode_uploaded_location(value: str) -> tuple[float, float] | None:
    """Resolve an actual uploaded location for the downloadable map image."""
    query = str(value).strip()
    if not query:
        return None
    if query in _GEOCODE_CACHE:
        return _GEOCODE_CACHE[query]
    try:
        url = "https://nominatim.openstreetmap.org/search?" + urlencode({
            "format": "jsonv2", "limit": 1, "q": query,
        })
        request_obj = Request(url, headers={"User-Agent": "FinSightAI/1.0"})
        with urlopen(request_obj, timeout=5) as response:
            results = json.loads(response.read().decode("utf-8"))
        if results:
            point = (float(results[0]["lat"]), float(results[0]["lon"]))
            _GEOCODE_CACHE[query] = point
            return point
    except (HTTPError, URLError, TimeoutError, ValueError, KeyError, OSError, json.JSONDecodeError):
        pass
    _GEOCODE_CACHE[query] = None
    return None


def _visualization_png(visual: dict[str, object], chart_type: str) -> io.BytesIO:
    """Render one current-dataset visualization into an in-memory PNG."""
    chart_type = str(chart_type).casefold()
    fig, axis = plt.subplots(figsize=(10, 5), constrained_layout=True)
    if chart_type == "donut":
        data = visual.get("categories") or []
        if not data:
            raise ValueError(visual.get("donut_message") or "No donut chart data is available.")
        axis.pie([float(row["value"]) for row in data],
                 labels=[str(row["category"]) for row in data],
                 autopct="%1.1f%%", startangle=90,
                 wedgeprops={"width": 0.42, "edgecolor": "white"})
        axis.set_title(f"{visual.get('category_metric_label', 'Value')} by {visual.get('category_label', 'Category')}")
    elif chart_type == "column":
        data = visual.get("column_data") or []
        if not data:
            raise ValueError(visual.get("column_message") or "No column chart data is available.")
        labels = [str(row["category"]) for row in data]
        axis.bar(labels, [float(row["value"]) for row in data], color="#2563eb")
        axis.set_title(f"{visual.get('category_metric_label', 'Value')} by {visual.get('category_label', 'Category')}")
        axis.set_xlabel(str(visual.get("category_label", "Category")))
        axis.set_ylabel(str(visual.get("category_metric_label", "Value")))
        axis.tick_params(axis="x", rotation=35)
    elif chart_type == "area":
        data = visual.get("area_data") or []
        if not data:
            raise ValueError(visual.get("area_message") or "No area chart data is available.")
        x_values = np.arange(len(data))
        periods = [str(row["period"]) for row in data]
        y_values = [float(row["value"]) for row in data]
        axis.fill_between(x_values, y_values, color="#69d6a3", alpha=0.35)
        axis.plot(x_values, y_values, color="#168b5a", linewidth=2)
        axis.set_title(f"{visual.get('metric_column', 'Metric')} over time")
        axis.set_xlabel(str(visual.get("date_column") or "Date"))
        axis.set_ylabel(str(visual.get("metric_column") or "Value"))
        axis.set_xticks(x_values)
        axis.set_xticklabels(periods)
        axis.tick_params(axis="x", rotation=35)
    elif chart_type == "totals":
        totals = visual.get("totals") or {}
        if not totals:
            raise ValueError("No numeric data is available for a totals chart.")
        axis.bar([str(key).replace("_", " ").title() for key in totals],
                 [float(value) for value in totals.values()], color="#16a34a")
        axis.set_title("Uploaded numeric totals")
        axis.tick_params(axis="x", rotation=35)
    elif chart_type == "distribution":
        data = visual.get("distribution_data") or []
        if not data:
            raise ValueError(visual.get("distribution_message") or "No numeric distribution is available.")
        axis.bar([str(row["range"]) for row in data],
                 [int(row["value"]) for row in data], color="#8cbcff")
        axis.set_title(f"Distribution of {visual.get('distribution_label', 'Value')}")
        axis.set_xlabel(str(visual.get("distribution_label", "Value")))
        axis.set_ylabel("Rows")
        axis.tick_params(axis="x", rotation=35)
    elif chart_type == "map":
        locations = visual.get("locations") or []
        if not locations:
            raise ValueError(visual.get("map_message") or "No geographic data available for map visualization.")
        resolved = []
        for row in locations[:20]:
            point = _geocode_uploaded_location(str(row["location"]))
            if point is not None:
                resolved.append((row, point))
        if not resolved:
            raise ValueError("The uploaded geographic values could not be resolved for a downloadable map.")
        values = [float(row["value"]) for row, _ in resolved]
        scatter = axis.scatter([point[1] for _, point in resolved],
                               [point[0] for _, point in resolved],
                               c=values, s=80, cmap="viridis", alpha=0.85,
                               edgecolors="black", linewidths=0.4)
        for row, point in resolved:
            axis.annotate(str(row["location"]), (point[1], point[0]),
                          xytext=(4, 4), textcoords="offset points", fontsize=8)
        axis.set_xlim(-180, 180)
        axis.set_ylim(-90, 90)
        axis.set_xlabel("Longitude")
        axis.set_ylabel("Latitude")
        axis.set_title(f"{visual.get('metric_column', 'Value')} by uploaded geographic value")
        fig.colorbar(scatter, ax=axis, label=str(visual.get("metric_column") or "Value"))
        axis.grid(alpha=0.25)
    else:
        plt.close(fig)
        raise ValueError("Unknown visualization.")
    axis.grid(alpha=0.2)
    output = io.BytesIO()
    fig.savefig(output, format="png", dpi=120)
    plt.close(fig)
    output.seek(0)
    return output


@app.get("/analytics/download/<chart_type>")
@login_required
def download_analytics_visualization(chart_type: str):
    """Download a PNG rendered from the authenticated user's active upload."""
    if chart_type not in {"donut", "column", "area", "map", "totals", "distribution"}:
        return jsonify({"error": "Unknown visualization."}), 404
    user_id = _session_user_id()
    _, df = _active_dataset_and_frame(user_id)
    if df is None or df.empty:
        return jsonify({"error": "Upload a dataset before downloading a visualization."}), 404
    try:
        visual = _visualization_data(df)
        output = _visualization_png(visual, chart_type)
    except (ValueError, TypeError, OSError) as exc:
        return jsonify({"error": str(exc) or "This visualization is not available for the current dataset."}), 404
    return send_file(output, as_attachment=True,
                     download_name=f"finsight_{chart_type}_chart.png", mimetype="image/png")


# =====================================================
# Routes - File upload
# =====================================================
@app.post("/upload/preview")
@login_required
def upload_preview():
    """Inspect an upload without persisting it or changing the active dataset."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided."}), 400
    file = request.files["file"]
    filename = (file.filename or "").strip()
    if not filename:
        return jsonify({"error": "No file selected."}), 400
    if not allowed_file(filename):
        return jsonify({"error": "File must be CSV, XLSX, XLS, or JSON."}), 400
    contents = file.stream.read()
    if not contents:
        return jsonify({"error": "File is empty."}), 400
    if len(contents) > MAX_CONTENT_LENGTH:
        limit = round(MAX_CONTENT_LENGTH / (1024 * 1024), 2)
        return jsonify({"error": f"Uploaded file is too large. Maximum size is {limit:g} MB."}), 413
    try:
        frame = _read_upload_dataframe(io.BytesIO(contents), filename)
        detection = detect_columns(frame)
        # Only semantic matches (or an explicit user mapping) are applied.
        # A random numeric/date column must never be relabeled as a financial
        # field merely to make forecasting appear available.
        preview_mapping = detection.get("mapping", {})
        profile = profile_dataframe(frame, preview_mapping)
        return jsonify({"success": True, "profile": profile})
    except (ValueError, pd.errors.ParserError, OSError) as exc:
        return jsonify({"error": str(exc) or "We could not read this file."}), 400
    except Exception:
        app.logger.exception("Upload preview failed")
        return jsonify({"error": "We could not read this file. Check its format and try again."}), 400


@app.route("/upload", methods=["GET", "POST"])
@login_required
def upload_file():
    """Accept CSV, XLSX, XLS, or JSON, map business fields, and persist the result."""
    user_id = _session_user_id()
    company_id = _session_company_id()

    if request.method == "GET":
        return render_template(
            "upload.html", company_name=session["company_name"], field_specs=FIELD_SPECS,
            max_upload_mb=round(MAX_CONTENT_LENGTH / (1024 * 1024), 2),
        )

    if "file" not in request.files:
        return jsonify({"error": "No file provided."}), 400
    file = request.files["file"]
    original_name = (file.filename or "").strip()
    if not original_name:
        return jsonify({"error": "No file selected."}), 400
    if not allowed_file(original_name):
        return jsonify({"error": "File must be CSV, XLSX, XLS, or JSON."}), 400
    safe_name = secure_filename(original_name)
    if not safe_name:
        return jsonify({"error": "The file name is not valid."}), 400

    resource = _powerbi_resource(user_id)
    user_upload_dir = POWERBI_ROOT / resource["folder_token"] / "uploads"
    user_upload_dir.mkdir(parents=True, exist_ok=True)
    stored_name = f"{uuid.uuid4().hex}_{safe_name}"
    disk_path = user_upload_dir / stored_name
    try:
        file.save(disk_path)
    except Exception:
        app.logger.exception("Could not save uploaded file")
        return jsonify({"error": "Could not save the file. Please try again."}), 500

    file_id: int | None = None
    upload_succeeded = False
    try:
        try:
            raw_df = _read_upload_dataframe(disk_path, safe_name)
        except Exception as exc:
            add_history(user_id, company_id, "upload", safe_name, status="failed",
                        details=f"read error: {exc}")
            return jsonify({"error": str(exc) or "We could not read this file. Check its format and try again."}), 400

        mapping = _mapping_from_request()
        detection = detect_columns(raw_df)
        if not mapping:
            mapping = detection.get("mapping", {})
        cleaned_df, cleaning = clean_dataframe(raw_df, mapping)
        mapped_labels = {field: mapping.get(field) for field in FIELD_SPECS if mapping.get(field)}
        warnings = list(detection.get("warnings", []))
        recommended_missing = [
            FIELD_SPECS[field]["label"] for field in ("tx_date", "revenue")
            if field not in mapping
        ]
        if recommended_missing:
            warnings.append(
                "Forecasting is optional for this upload. Recommended forecasting "
                "fields not detected: " + ", ".join(recommended_missing) +
                ". All uploaded columns were retained for cleaning and analysis."
            )

        file_id = run_query(
            """INSERT INTO uploaded_files
               (user_id, company_id, version, original_name, stored_name, file_size,
                 rows_imported, status, source_format, source_columns, column_mapping,
                 cleaning_summary, upload_warnings, raw_data)
               VALUES (%s, %s, %s, %s, %s, %s, %s, 'processing', %s, %s, %s, %s, %s, %s)""",
            (user_id, company_id, _next_dataset_version(user_id), safe_name, stored_name,
             disk_path.stat().st_size, int(len(raw_df)), Path(safe_name).suffix.lower().lstrip("."),
             json.dumps([str(column) for column in raw_df.columns], ensure_ascii=False),
             json.dumps(mapped_labels, ensure_ascii=False), json.dumps(cleaning),
             json.dumps(warnings, ensure_ascii=False), disk_path.read_bytes()),
            commit=True,
        )["last_id"]

        # Keep the arbitrary cleaned frame as the durable source of truth.
        # Canonical nullable fields below are only a compatibility projection
        # for the legacy financial_data table and must not leak into downloads.
        dataset_frame = cleaned_df.copy()
        financial_frame = cleaned_df.copy()
        for column in CANONICAL_NUMERIC_FIELDS:
            if column not in financial_frame.columns:
                financial_frame[column] = np.nan
        for column in CANONICAL_TEXT_FIELDS:
            if column not in financial_frame.columns:
                financial_frame[column] = None
            else:
                financial_frame[column] = financial_frame[column].astype(object).where(financial_frame[column].notna(), None)
        if "tx_date" not in financial_frame.columns:
            financial_frame["tx_date"] = pd.NaT

        financial_sql = """INSERT INTO financial_data
            (user_id, company_id, uploaded_file_id, tx_date, transaction_id, description,
             amount, revenue, expenses, profit, customers, marketing_spend,
             tx_type, category, payment_method, department, city, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        financial_rows = []
        dataset_rows = []
        for row_number, (_, row) in enumerate(dataset_frame.iterrows(), start=1):
            financial_row = financial_frame.iloc[row_number - 1]
            tx_value = financial_row.get("tx_date")
            tx_value = pd.to_datetime(tx_value, errors="coerce") if tx_value is not None else pd.NaT
            financial_rows.append((
                user_id, company_id, file_id,
                None if pd.isna(tx_value) else tx_value.date(),
                _db_text(financial_row.get("transaction_id")), _db_text(financial_row.get("description")),
                *[None if pd.isna(financial_row.get(column)) else float(financial_row.get(column))
                  for column in ("amount", "revenue", "expenses", "profit", "customers", "marketing_spend")],
                _db_text(financial_row.get("tx_type")), _db_text(financial_row.get("category")), _db_text(financial_row.get("payment_method")),
                _db_text(financial_row.get("department")), _db_text(financial_row.get("city")), _db_text(financial_row.get("status")),
            ))
            dataset_rows.append((
                user_id, company_id, file_id, row_number,
                json.dumps(json_safe_record(row.to_dict()), ensure_ascii=False, allow_nan=False),
            ))

        conn = get_db()
        try:
            cur = conn.cursor()
            for start in range(0, len(financial_rows), 500):
                cur.executemany(financial_sql, financial_rows[start:start + 500])
                cur.executemany(
                    """INSERT INTO dataset_rows
                       (user_id, company_id, uploaded_file_id, `row_number`, row_data)
                       VALUES (%s, %s, %s, %s, %s)""",
                    dataset_rows[start:start + 500],
                )
            conn.commit()
            cur.close()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

        run_query(
            """UPDATE uploaded_files SET rows_imported=%s, status='processed',
               source_columns=%s, column_mapping=%s, cleaning_summary=%s, upload_warnings=%s
               WHERE id=%s AND user_id=%s""",
             (int(len(dataset_frame)), json.dumps([str(column) for column in raw_df.columns], ensure_ascii=False),
             json.dumps(mapped_labels, ensure_ascii=False), json.dumps(cleaning),
             json.dumps(warnings, ensure_ascii=False), file_id, user_id), commit=True,
        )

        train_df = dataset_frame.copy()
        if "tx_date" in train_df.columns:
            train_df["tx_date"] = pd.to_datetime(train_df["tx_date"], errors="coerce")
            train_df = train_df.dropna(subset=["tx_date"])
        else:
            train_df = pd.DataFrame()
        # Keep all cleaned rows visible to the application. The training frame
        # is allowed to exclude invalid dates, but that modeling choice must
        # never change dashboard counts or the cleaned download.
        active_frame = dataset_frame.copy()
        session_data[user_id] = {
            "df": active_frame,
            "dataset_id": file_id,
            "models": {"amount": None, "expenses": None, "revenue": None, "profit": None, "risk": None},
        }
        if not train_df.empty:
            train_df["Date_Number"] = (train_df["tx_date"] - train_df["tx_date"].min()).dt.days
            train_df["Month"] = train_df["tx_date"].dt.month
            train_df["Day_of_Week"] = train_df["tx_date"].dt.dayofweek
            try:
                _train_models_for(user_id, train_df)
            except Exception as exc:
                app.logger.warning("Model training skipped: %s", exc)

        training_targets = [
            column for column in ("amount", "expenses", "revenue", "profit")
            if column in train_df.columns
            and pd.to_numeric(train_df[column], errors="coerce").notna().sum() >= 10
        ]
        forecasting_available = bool(len(train_df) >= 10 and training_targets)
        forecasting_message = (
            "Forecasting is available for the detected historical numeric fields."
            if forecasting_available else
            "Forecasting was not generated because this upload does not yet contain "
            "at least 10 valid dated observations for a supported numeric measure. "
            "Cleaning, analysis, and exports remain available."
        )

        _run_universal_analysis(user_id, dataset_frame)
        add_history(user_id, company_id, "upload", safe_name, file_id=file_id,
                     status="processed", details=f"{len(dataset_frame)} rows imported")
        try:
            _generate_powerbi_resources(user_id, company_id, session["company_name"], file_id)
        except Exception as exc:
            # Power BI artefacts are optional and must not invalidate a valid upload.
            app.logger.warning("Power BI resource generation skipped: %s", exc)

        stats = {
            "rows": int(len(dataset_frame)),
            "rows_detected": int(len(raw_df)),
            "columns": int(len(raw_df.columns)),
            "duplicates_removed": cleaning["duplicates_removed"],
            "blank_rows_removed": cleaning["blank_rows_removed"],
            "missing_values_remaining": cleaning["missing_values_remaining"],
            "missing_values_filled": cleaning.get("missing_values_filled", 0),
            "invalid_dates": cleaning["invalid_dates"],
            "date_column": mapping.get("tx_date"),
            "mapping": mapped_labels,
            "warnings": warnings,
            "missing_recommended": recommended_missing,
            "forecasting_available": forecasting_available,
            "forecasting_message": forecasting_message,
            "preview": [json_safe_record(record) for record in dataset_frame.head(8).to_dict(orient="records")],
        }
        for column in ("revenue", "expenses", "profit", "amount", "customers", "marketing_spend"):
            if column in dataset_frame.columns and dataset_frame[column].notna().any():
                stats[f"{column}_total"] = float(pd.to_numeric(dataset_frame[column], errors="coerce").sum())
        detected_types = _analysis_types_for(dataset_frame)
        date_columns = [column for column, kind in detected_types.items() if kind == "date"]
        valid_dates = pd.Series(dtype="datetime64[ns]")
        if date_columns:
            valid_dates = pd.to_datetime(dataset_frame[date_columns[0]], errors="coerce").dropna()
        stats["date_range"] = (f"{valid_dates.min().date()} to {valid_dates.max().date()}"
                                if not valid_dates.empty else "")
        stats["numeric_columns"] = [column for column, kind in detected_types.items() if kind == "numeric"]
        stats["categorical_columns"] = [column for column, kind in detected_types.items() if kind in {"categorical", "text", "boolean"}]
        stats["detected_fields"] = detection["fields"]

        upload_succeeded = True
        return jsonify({"success": True, "stats": stats, "file_id": file_id})
    except mysql.connector.Error:
        app.logger.exception("Database failure while processing upload for user %s", user_id)
        if file_id is not None:
            try:
                run_query("UPDATE uploaded_files SET status='failed' WHERE id=%s AND user_id=%s",
                          (file_id, user_id), commit=True)
            except Exception:
                app.logger.exception("Could not mark database-failed upload %s as failed", file_id)
        return jsonify({"error": "The dataset could not be saved because the database is unavailable. Please try again later."}), 503
    except ValueError as exc:
        app.logger.info("Invalid upload from user %s: %s", user_id, exc)
        if file_id is not None:
            try:
                run_query("UPDATE uploaded_files SET status='failed' WHERE id=%s AND user_id=%s",
                          (file_id, user_id), commit=True)
            except Exception:
                app.logger.exception("Could not mark invalid upload %s as failed", file_id)
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        app.logger.exception("Upload processing failed for user %s", user_id)
        if file_id is not None:
            try:
                run_query("UPDATE uploaded_files SET status='failed' WHERE id=%s AND user_id=%s",
                          (file_id, user_id), commit=True)
            except Exception:
                app.logger.exception("Could not mark failed upload %s as failed", file_id)
        try:
            add_history(user_id, company_id, "upload", safe_name, file_id=file_id,
                        status="failed", details=str(exc))
        except Exception:
            app.logger.exception("Could not record failed upload")
        return jsonify({"error": "We could not process this file. Please try again."}), 500
    finally:
        if not upload_succeeded and disk_path.exists():
            try:
                disk_path.unlink()
            except OSError:
                app.logger.warning("Could not remove failed upload %s", disk_path)


def _train_models_for(user_id: int, df: pd.DataFrame) -> None:
    """Train regression models using improved ML pipeline.
    
    Uses:
    - Chronological train/test split (not random, for time-series)
    - Feature engineering (temporal, lag, rolling, derived features)
    - Multiple model comparison (Linear, Ridge, Lasso, Tree, Forest, Boosting)
    - Cross-validation (TimeSeriesSplit) for stability assessment
    - Comprehensive evaluation metrics (MAE, RMSE, MSE, R², overfitting detection)
    - Model persistence to disk with metadata
    
    Still maintains session-based models for backward compatibility.
    """
    if len(df) < 10:
        models = get_models(user_id)
        for key in ("amount", "expenses", "revenue", "profit"):
            models[key] = None
        return

    models = get_models(user_id)
    
    # Use new ML pipeline if available
    if get_pipeline is not None:
        try:
            pipeline = get_pipeline()
            
            # Train all regression models using the pipeline
            for target in ["amount", "revenue", "expenses", "profit"]:
                if target not in df.columns:
                    models[target] = None
                    continue
                
                result = pipeline.train_regression_model(df, target, user_id, model_type='best')
                
                if result.get('success'):
                    best_model_name = result.get('selected_model', 'linear')
                    metrics = result.get('metrics', {})
                    
                    # Store in session for immediate use
                    try:
                        loaded_model, metadata = pipeline._load_model(user_id, target)
                        models[target] = {
                            "model": loaded_model,
                            "df_min": df["tx_date"].min() if "tx_date" in df.columns else None,
                            "metrics": {
                                "mae": float(metrics.get("test_mae", 0)),
                                "r2": float(metrics.get("test_r2", 0)),
                                "rmse": float(metrics.get("test_rmse", 0)),
                                "train_r2": float(metrics.get("train_r2", 0)),
                                "cv_r2_mean": float(metrics.get("cv_r2_mean", 0)),
                                "overfitting": metrics.get("overfitting_warning", "Unknown"),
                                "model_type": best_model_name,
                                "features_used": len(metrics.get("features_used", [])),
                                "train_samples": metrics.get("n_train", 0),
                                "test_samples": metrics.get("n_test", 0),
                            }
                        }
                        print(f"[finsight] Trained {target} model ({best_model_name}): "
                              f"R²={metrics.get('test_r2', 0):.4f}, "
                              f"MAE={metrics.get('test_mae', 0):.2f}, "
                              f"Overfitting={metrics.get('overfitting_warning', 'unknown')}")
                    except Exception as e:
                        print(f"[finsight] Could not load trained {target} model: {e}")
                        models[target] = None
                else:
                    error = result.get('error', 'Unknown error')
                    print(f"[finsight] Could not train {target} model: {error}")
                    models[target] = None
            
            _train_risk_model_for(user_id, df)
            return
        except Exception as e:
            print(f"[finsight] ML pipeline error: {e}")
            traceback.print_exc()
    
    # Fallback to old training if pipeline not available
    print("[finsight] Falling back to legacy model training")
    if LinearRegression is None or train_test_split is None:
        app.logger.warning("Regression training skipped because scikit-learn is unavailable.")
        return
    feature_cols = ["Date_Number", "Month", "Day_of_Week"]
    for key, target in (("amount", "amount"),
                        ("expenses", "expenses"),
                        ("revenue", "revenue"),
                        ("profit", "profit")):
        if target not in df.columns:
            continue
        valid = pd.to_numeric(df[target], errors="coerce").notna()
        if not valid.any():
            continue
        target_frame = df.loc[valid, feature_cols + [target]].copy()
        X = target_frame[feature_cols].apply(pd.to_numeric, errors="coerce")
        y = pd.to_numeric(target_frame[target], errors="coerce")
        valid_features = X.notna().all(axis=1) & y.notna()
        X = X.loc[valid_features]
        y = y.loc[valid_features]
        if len(y) < 2:
            continue
        if y.nunique() < 2:
            continue
        try:
            x_train, x_test, y_train, y_test = train_test_split(
                X, y, test_size=0.2, random_state=42,
            )
            m = LinearRegression().fit(x_train, y_train)
            models[key] = {"model": m, "df_min": df["tx_date"].min(),
                           "metrics": {
                               "mae": float(mean_absolute_error(y_test, m.predict(x_test))),
                               "r2":  float(r2_score(y_test, m.predict(x_test))),
                           }}
        except Exception as exc:
            print(f"[finsight] Could not train {key} model: {exc}")
            models[key] = None
    
    # Also train risk model in fallback
    _train_risk_model_for(user_id, df)



def _train_risk_model_for(user_id: int, df: pd.DataFrame) -> None:
    """Train risk classifier using improved ML pipeline.
    
    Uses:
    - Chronological daily aggregation
    - Pre-defined risk thresholds with no target leakage
    - Transparent rule-based risk scoring
    - Model persistence to disk with metadata
    
    Risk classification is based on deterministic business rules:
    - HIGH RISK: Negative profit OR expenses > revenue
    - MEDIUM RISK: Profit margin < 8% OR expense ratio > 85%
    - LOW RISK: Otherwise
    
    These rules are applied to create labels BEFORE any features are used.
    """
    models = get_models(user_id)
    required_features = ("revenue", "expenses", "profit", "amount")
    if (
        df is None or df.empty or "tx_date" not in df.columns
        or any(column not in df.columns for column in required_features)
    ):
        models["risk"] = None
        models["risk_error"] = (
            "Risk classification is unavailable because this dataset does not "
            "contain four sufficiently populated dated measures: revenue, "
            "expenses, profit, and amount."
        )
        return

    working = df.copy()
    working["tx_date"] = pd.to_datetime(working["tx_date"], errors="coerce")
    for column in required_features:
        working[column] = pd.to_numeric(working[column], errors="coerce")
    complete = working.dropna(subset=["tx_date", *required_features]).copy()
    if len(complete) < MIN_FORECAST_OBSERVATIONS:
        models["risk"] = None
        models["risk_error"] = (
            "Risk classification is unavailable because this dataset does not "
            f"contain {MIN_FORECAST_OBSERVATIONS} complete dated rows for revenue, "
            "expenses, profit, and amount."
        )
        return
    df = complete
    
    # Use new ML pipeline if available
    if get_pipeline is not None:
        try:
            pipeline = get_pipeline()
            result = pipeline.train_risk_classifier(df, user_id)
            
            if result.get('success'):
                metrics = result.get('metrics', {})
                
                # Store in session for immediate use
                try:
                    loaded_model, metadata = pipeline._load_model(user_id, 'risk')
                    features = metadata.get('features', ['revenue', 'expenses', 'profit', 'amount'])
                    models["risk"] = {
                        "model": loaded_model,
                        "features": features,
                        "periods": metrics.get('n_daily_periods', 0),
                        "metrics": {
                            "train_accuracy": metrics.get("train_accuracy", 0),
                            "test_accuracy": metrics.get("test_accuracy", 0),
                            "test_f1": metrics.get("test_f1_weighted", 0),
                            "overfitting": metrics.get("overfitting_warning", "Unknown"),
                        }
                    }
                    models.pop("risk_error", None)
                    print(f"[finsight] Trained risk classifier: "
                          f"accuracy={metrics.get('test_accuracy', 0):.4f}, "
                          f"F1={metrics.get('test_f1_weighted', 0):.4f}, "
                          f"periods={metrics.get('n_daily_periods', 0)}")
                except Exception as e:
                    print(f"[finsight] Could not load trained risk model: {e}")
                    models["risk"] = None
            else:
                error = result.get('error', 'Unknown error')
                print(f"[finsight] Could not train risk classifier: {error}")
                models["risk"] = None
                models["risk_error"] = error
            
            return
        except Exception as e:
            print(f"[finsight] Risk ML pipeline error: {e}")
            traceback.print_exc()
    
    # If the ML module is unavailable, retain the same transparent risk rules
    # in session rather than falling back to a classifier trained on its own
    # derived labels.
    print("[finsight] Falling back to rule-based risk scoring")
    features = ["revenue", "expenses", "profit", "amount"]
    if (RuleBasedRiskModel is None or df.empty or "tx_date" not in df.columns
            or any(c not in df.columns for c in features)):
        models["risk"] = None
        return
    working = df.copy()
    periods = working.groupby(working["tx_date"].dt.date, as_index=False)[features].sum()
    if periods.empty:
        models["risk"] = None
        return
    try:
        rule_model = RuleBasedRiskModel()
        models["risk"] = {"model": rule_model, "features": features, "periods": len(periods)}
    except Exception as exc:
        print(f"[finsight] Could not train risk classifier: {exc}")
        models["risk"] = None


# =====================================================
# Routes - Predictions
# =====================================================
MIN_FORECAST_OBSERVATIONS = 12


def _analysis_types_for(df: pd.DataFrame | None, *, use_cached: bool = True) -> dict[str, str]:
    if use_cached:
        analysis = session_data.get(_session_user_id() or -1, {}).get("analysis", {})
        types = analysis.get("types") if isinstance(analysis, dict) else None
        if types:
            return types
    if universal_analysis is not None and df is not None:
        return universal_analysis.detect_types(df)
    # Keep structural analytics available when an optional native ML wheel is
    # unavailable. Prediction remains disabled in that case, but numeric and
    # categorical charts can still be derived from the real upload.
    profile = detect_columns(df) if df is not None else {"columns": []}
    type_map = {
        "numeric": "numeric", "date": "date", "categorical": "categorical",
        "text": "text", "empty": "empty", "constant": "constant",
    }
    return {
        str(item.get("source")): type_map.get(item.get("type"), "other")
        for item in profile.get("columns", []) if item.get("source") is not None
    }


def _date_column_for(df: pd.DataFrame | None, types: dict[str, str] | None = None) -> str | None:
    if df is None or df.empty:
        return None
    types = types or _analysis_types_for(df)
    return next((column for column, kind in types.items()
                 if kind == "date" and column in df.columns), None)


def _forecast_targets(df: pd.DataFrame | None) -> list[str]:
    """Return actual numeric columns with a valid dated history and model."""
    if df is None or df.empty or universal_analysis is None:
        return []
    types = _analysis_types_for(df)
    date_column = _date_column_for(df, types)
    if not date_column:
        return []
    dates = pd.to_datetime(df[date_column], errors="coerce")
    targets = []
    for section in session_data.get(_session_user_id() or -1, {}).get("analysis", {}).get("sections", []):
        target = section.get("target")
        if (section.get("kind") != "regression" or section.get("error")
                or section.get("prediction_value") is None or target not in df.columns):
            continue
        values = pd.to_numeric(df[target], errors="coerce")
        valid_values = values[dates.notna()].dropna()
        if (len(valid_values) >= MIN_FORECAST_OBSERVATIONS and valid_values.nunique() >= 2):
            targets.append(str(target))
    return targets


def _risk_data_ready(df: pd.DataFrame | None) -> bool:
    """Return whether the legacy risk view has enough complete real rows."""
    required = ("tx_date", "revenue", "expenses", "profit", "amount")
    if df is None or df.empty or any(column not in df.columns for column in required):
        return False
    work = df[list(required)].copy()
    work["tx_date"] = pd.to_datetime(work["tx_date"], errors="coerce")
    for column in required[1:]:
        work[column] = pd.to_numeric(work[column], errors="coerce")
    return len(work.dropna()) >= MIN_FORECAST_OBSERVATIONS


def _forecast_step_days(df: pd.DataFrame, date_column: str | None = None) -> int:
    date_column = date_column or _date_column_for(df)
    if not date_column or date_column not in df.columns:
        return 1
    dates = pd.to_datetime(df[date_column], errors="coerce").dropna().sort_values().drop_duplicates()
    if len(dates) < 2:
        return 1
    step = dates.diff().dt.total_seconds().div(86400).dropna().median()
    try:
        return max(1, min(366, int(round(float(step)))))
    except (TypeError, ValueError):
        return 1


def _forecast_status(user_id: int, df: pd.DataFrame | None) -> dict[str, object]:
    """Explain forecast readiness without turning data quality into a crash."""
    status: dict[str, object] = {
        "valid_dates": 0,
        "invalid_dates": 0,
        "date_column_found": False,
        "date_mapping_found": True,
        "target_counts": {},
        "warning": "",
    }
    if df is None or df.empty:
        status["warning"] = "Upload a dataset to enable forecasting."
        return status

    types = _analysis_types_for(df)
    date_column = _date_column_for(df, types)
    if not date_column:
        status["warning"] = (
            "Forecasting is currently unavailable because no usable date column was found. "
            "Your data is still available for dashboard and analytics use."
        )
        return status

    dates = pd.to_datetime(df[date_column], errors="coerce")
    valid_date_count = int(dates.notna().sum())
    status["valid_dates"] = valid_date_count
    status["date_column_found"] = bool(valid_date_count)
    try:
        upload = run_query(
            """SELECT column_mapping, cleaning_summary FROM uploaded_files
               WHERE user_id=%s AND status='processed' ORDER BY id DESC LIMIT 1""",
            (user_id,), fetchone=True,
        )["row"] or {}
        cleaning = json.loads(upload.get("cleaning_summary") or "{}")
        status["invalid_dates"] = int(cleaning.get("invalid_dates") or 0)
    except Exception as exc:
        # The data itself remains authoritative; quality metadata is optional.
        app.logger.warning("Forecast quality details unavailable: %s", exc)

    counts: dict[str, int] = {}
    for target, kind in types.items():
        if kind != "numeric" or target not in df.columns:
            continue
        counts[target] = int(pd.to_numeric(df.loc[dates.notna(), target], errors="coerce").notna().sum())
    status["target_counts"] = counts
    if not status["date_column_found"]:
        invalid_note = (
            f" {status['invalid_dates']} date values could not be parsed."
            if status["invalid_dates"] else ""
        )
        status["warning"] = (
            "Forecasting is currently unavailable because no valid dated observations were found."
            f"{invalid_note} Check the date column and supported date formats. Your data is still "
            "available for dashboard and analytics use."
        )
    else:
        best_count = max(counts.values(), default=0)
        if best_count < MIN_FORECAST_OBSERVATIONS:
            invalid_note = (
                f" {status['invalid_dates']} date values could not be parsed."
                if status["invalid_dates"] else ""
            )
            status["warning"] = (
                f"Forecasting is currently unavailable because the uploaded dataset contains only "
                f"{best_count} valid dated numeric observations; at least {MIN_FORECAST_OBSERVATIONS} "
                f"are needed for a reliable forecast.{invalid_note} Your data has still been successfully "
                "loaded and can be used in the dashboard and analytics."
            )
        elif not _forecast_targets(df):
            status["warning"] = (
                "Forecasting is currently unavailable because no actual numeric column has enough "
                "dated observations and variation for a reliable model. Your data is still available "
                "for dashboard and analytics use."
            )
    return status


@app.route("/predict", methods=["GET", "POST"])
@login_required
def predict():
    user_id = _session_user_id()
    company_id = _session_company_id()

    if request.method == "GET":
        _, df = _active_dataset_and_frame(user_id)
        analysis = _analysis_context(user_id)
        analysis_types = (analysis or {}).get("types", {})
        date_column = _date_column_for(df, analysis_types)
        available_dates = []
        if df is not None and not df.empty and date_column:
            available_dates = sorted(
                pd.to_datetime(df[date_column], errors="coerce")
                .dropna().dt.strftime("%Y-%m-%d").unique().tolist()
            )
        forecast_targets = _forecast_targets(df)
        forecast_status = _forecast_status(user_id, df)
        if available_dates:
            next_history_date = (pd.to_datetime(available_dates[-1]) +
                                 pd.Timedelta(days=_forecast_step_days(df, date_column))).date()
            default_date = max(date.today() + timedelta(days=1), next_history_date).isoformat()
        else:
            default_date = (date.today() + timedelta(days=1)).isoformat()
        # Universal, dataset-agnostic prediction sections + business insight.
        return render_template(
            "predict.html",
            company_name=session["company_name"],
            has_data=bool(df is not None and not df.empty),
            risk_min_date=(date.today() + timedelta(days=1)).isoformat(),
            risk_max_date="",
            risk_default_date=default_date,
            forecast_targets=forecast_targets,
            expense_targets=[target for target in forecast_targets if target != (
                "revenue" if "revenue" in forecast_targets else (forecast_targets[0] if forecast_targets else None)
            )],
            risk_available=all(
                target in forecast_targets for target in ("revenue", "expenses", "profit", "amount")
            ) and _risk_data_ready(df),
            analysis=analysis if analysis is not None else {},
            insight=(analysis or {}).get("insight", ""),
            section_types=(analysis or {}).get("types", {}),
            date_column=date_column or "—",
            forecast_status=forecast_status,
        )

    if not request.is_json:
        return jsonify({"error": _message("api.error.predict_json_required")}), 400
    payload = request.get_json(silent=True) or {}
    model_type = payload.get("model_type")
    if not isinstance(model_type, str) or not model_type.strip():
        return jsonify({"error": "Choose a detected numeric column before forecasting."}), 400
    model_type = model_type.strip()
    date_str = payload.get("date")
    try:
        forecast_periods = max(1, min(24, int(payload.get("forecast_periods", 1))))
    except (TypeError, ValueError):
        return jsonify({"error": "Forecast horizon must be a whole number between 1 and 24."}), 400

    try:
        route_dataset_id, route_df = _active_dataset_and_frame(user_id)
    except mysql.connector.Error:
        app.logger.exception("Database failure while loading prediction dataset for user %s", user_id)
        return jsonify({"error": "Forecasting is temporarily unavailable because the database could not be reached."}), 503
    route_analysis = _analysis_context(user_id) or {}
    route_types = route_analysis.get("types", {})
    route_date_column = _date_column_for(route_df, route_types)
    route_section = next(
        (section for section in session_data.get(user_id, {}).get("analysis", {}).get("sections", [])
         if section.get("target") == model_type),
        None,
    )
    # A column named like a legacy financial target is still generic when the
    # uploaded dataset uses a different date column. Route it through the
    # universal model rather than requiring the old financial schema.
    use_dynamic_prediction = model_type not in ("amount", "expenses", "revenue", "profit") or (
        route_section is not None and route_date_column not in {None, "tx_date"}
    )

    # ---------- Dynamic prediction support for arbitrary uploaded columns ----------
    if use_dynamic_prediction:
        dataset_id, df = route_dataset_id, route_df
        analysis = _analysis_context(user_id)
        types = (analysis or {}).get("types", {})
        date_column = _date_column_for(df, types)
        model = next(
            (s for s in session_data.get(user_id, {}).get("analysis", {}).get("sections", [])
             if s.get("target") == model_type), None)
        if (dataset_id is None or df is None or df.empty or universal_analysis is None
                or not date_column or types.get(model_type) != "numeric" or model is None
                or model.get("error") or model.get("prediction_value") is None):
            return jsonify({
                "error": "This uploaded dataset does not contain a usable dated numeric target for forecasting."
            }), 400
        if not date_str:
            return jsonify({"error": _message("api.error.predict_date_required")}), 400
        try:
            target_date = pd.to_datetime(date_str).date()
        except Exception:
            return jsonify({"error": _message("api.error.predict_invalid_date")}), 400
        if len(str(model_type)) > 255:
            return jsonify({"error": "The selected column name is too long to store a prediction."}), 400
        result = universal_analysis.predict_dynamic_periods(
            model, forecast_periods, start_date=target_date
        )
        if not result.get("ok"):
            return jsonify({"error": result.get("error", "Prediction could not be generated.")}), 400
        step_days = _forecast_step_days(df, date_column)
        result_dates = result.get("dates") or [
            (target_date + timedelta(days=step_days * index)).isoformat()
            for index in range(forecast_periods)
        ]
        actual_dates = pd.to_datetime(df[date_column], errors="coerce").dt.date
        forecast_points = []
        for value, point_date in zip(result.get("values", []), result_dates):
            if not np.isfinite(float(value)):
                return jsonify({"error": "The model returned an invalid prediction for this dataset."}), 400
            parsed_date = pd.to_datetime(point_date, errors="coerce")
            actual_value = None
            if not pd.isna(parsed_date):
                matching = df[actual_dates == parsed_date.date()]
                if not matching.empty:
                    actual = pd.to_numeric(matching[model_type], errors="coerce").dropna()
                    if not actual.empty:
                        actual_value = round(float(actual.sum()), 2)
            forecast_points.append({
                "date": str(point_date), "value": round(float(value), 2), "actual": actual_value,
            })
        if not forecast_points:
            return jsonify({"error": "The model returned no predictions for this dataset."}), 400
        prediction_info = {
            "model_name": model.get("model_name", "Linear Regression"),
            "estimated_error": (model.get("metrics") or {}).get("rmse")
                or (model.get("metrics") or {}).get("mae") or 0,
            "metrics": model.get("metrics") or {},
        }
        prediction_ids = []
        try:
            for point in forecast_points:
                actual_value = point["actual"]
                prediction_id = run_query(
                    """INSERT INTO predictions
                       (user_id, company_id, uploaded_file_id, prediction_type, prediction_date,
                        actual_value, predicted_value, prediction_error, model_name)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (user_id, company_id, dataset_id, str(model_type), point["date"], actual_value,
                     point["value"], round(actual_value - point["value"], 2)
                     if actual_value is not None else None, prediction_info["model_name"]),
                    commit=True,
                )["last_id"]
                prediction_ids.append(prediction_id)
        except mysql.connector.Error:
            app.logger.exception("Database failure while saving dynamic prediction for user %s", user_id)
            return jsonify({"error": "Prediction could not be saved because the database is unavailable. Please try again later."}), 503
        except (TypeError, ValueError, OSError) as exc:
            app.logger.exception("Dynamic prediction persistence failed for user %s", user_id)
            return jsonify({"error": f"Prediction could not be saved for this dataset: {exc}"}), 422
        if not prediction_ids:
            return jsonify({"error": "Prediction could not be saved for this dataset."}), 422
        try:
            add_history(
                user_id, company_id, "prediction",
                f"{str(model_type).replace('_', ' ').title()} forecast from {target_date}",
                prediction_id=prediction_ids[0], file_id=dataset_id, status="ok",
                details=f"periods={forecast_periods}; first_value={forecast_points[0]['value']}",
            )
        except mysql.connector.Error:
            # A prediction has already been saved. Keep it usable and surface a
            # database status instead of converting the successful forecast to a
            # generic Flask 500 page.
            app.logger.exception("Database failure while recording dynamic prediction history")
            return jsonify({"error": "Prediction was generated, but its history could not be saved. Please try again later."}), 503
        except Exception:
            app.logger.exception("Could not record dynamic prediction history")
        try:
            _generate_powerbi_resources(user_id, company_id, session["company_name"], dataset_id)
        except Exception as exc:
            app.logger.warning("Power BI resource refresh skipped after prediction: %s", exc)
        try:
            chart = _build_chart(df, str(model_type), forecast_points, date_column=date_column)
        except Exception as chart_exc:
            app.logger.warning("Generic prediction chart rendering skipped: %s", chart_exc)
            chart = ""
        history = []
        for _, row in df.dropna(subset=[date_column, model_type]).tail(60).iterrows():
            parsed = pd.to_datetime(row[date_column], errors="coerce")
            value = pd.to_numeric(pd.Series([row[model_type]]), errors="coerce").iloc[0]
            if not pd.isna(parsed) and not pd.isna(value):
                history.append({"date": str(parsed.date()), "value": float(value)})
        return jsonify({
            "success": True, "prediction": forecast_points[0]["value"],
            "label": str(forecast_points[0]["value"]), "forecasts": forecast_points,
            "frequency": result.get("frequency", "period"), "date": str(target_date),
            "model_type": model_type, "kind": "regression", "chart": chart,
            "history": history, "prediction_id": prediction_ids[0],
            "prediction_ids": prediction_ids, "model": prediction_info,
        })
    if not date_str:
        return jsonify({"error": _message("api.error.predict_date_required")}), 400

    try:
        target_date = pd.to_datetime(date_str).date()
    except Exception:
        return jsonify({"error": _message("api.error.predict_invalid_date")}), 400

    try:
        dataset_id, df = _active_dataset_and_frame(user_id)
    except mysql.connector.Error:
        app.logger.exception("Database failure while loading legacy prediction dataset for user %s", user_id)
        return jsonify({"error": "Forecasting is temporarily unavailable because the database could not be reached."}), 503
    models = get_models(user_id)
    model_info = models.get(model_type)

    if dataset_id is None or df is None or df.empty:
        return jsonify({"error": _message("api.error.predict_need_data")}), 400
    if "tx_date" not in df.columns or pd.to_datetime(df["tx_date"], errors="coerce").notna().sum() == 0:
        return jsonify({
            "error": "Forecasting requires a usable date column in the uploaded dataset."
        }), 400
    if model_type not in df.columns or pd.to_numeric(df[model_type], errors="coerce").notna().sum() < MIN_FORECAST_OBSERVATIONS:
        return jsonify({
            "error": f"The uploaded dataset does not contain at least {MIN_FORECAST_OBSERVATIONS} valid observations for '{model_type}'."
        }), 400
    # Models from a previously uploaded dataset can be absent after a restart
    # or after a validation fix. Rebuild them from the active upload instead
    # of incorrectly asking the user to upload the same file again.
    if model_info is None:
        try:
            training_df = df.copy()
            training_df["tx_date"] = pd.to_datetime(training_df["tx_date"], errors="coerce")
            training_df = training_df.dropna(subset=["tx_date"])
            training_df["Date_Number"] = (training_df["tx_date"] - training_df["tx_date"].min()).dt.days
            training_df["Month"] = training_df["tx_date"].dt.month
            training_df["Day_of_Week"] = training_df["tx_date"].dt.dayofweek
            _train_models_for(user_id, training_df)
        except Exception:
            app.logger.exception("Could not rebuild legacy forecast models for user %s", user_id)
            return jsonify({"error": "Forecasting could not prepare a model for this dataset."}), 422
        models = get_models(user_id)
        model_info = models.get(model_type)
    if model_info is None:
        return jsonify({"error": _message("api.error.predict_need_data")}), 400

    try:
        step_days = _forecast_step_days(df)
        forecast_points = []
        prediction_info = {}
        for period in range(forecast_periods):
            forecast_date = target_date + timedelta(days=step_days * period)
            if get_pipeline is not None:
                value, info = get_pipeline().predict_regression(
                    user_id, model_type, {}, history_df=df, prediction_date=forecast_date
                )
                # A model may have been persisted by an older pipeline
                # version or for a previous upload. Rebuild once from the
                # active dataset before returning a user-facing failure.
                if value is None:
                    _train_models_for(user_id, df)
                    value, info = get_pipeline().predict_regression(
                        user_id, model_type, {}, history_df=df, prediction_date=forecast_date
                    )
                if value is None:
                    return jsonify({"error": info.get("error") or _message("api.error.predict_invalid_input")}), 400
                predicted_value = float(value)
                if not np.isfinite(predicted_value):
                    return jsonify({"error": "The model returned an invalid prediction for this dataset."}), 400
                prediction_info = info or {}
            else:
                date_number = (pd.to_datetime(forecast_date) - model_info["df_min"]).days
                X_future = pd.DataFrame({"Date_Number": [date_number], "Month": [forecast_date.month],
                                             "Day_of_Week": [forecast_date.weekday()]})
                predicted_value = float(model_info["model"].predict(X_future)[0])
                prediction_info = {
                    "model_name": "linear_regression",
                    "estimated_error": model_info.get("metrics", {}).get("rmse", 0),
                }
            if not np.isfinite(predicted_value):
                return jsonify({"error": "The model returned an invalid prediction for this dataset."}), 400
            matching = df[pd.to_datetime(df["tx_date"], errors="coerce").dt.date == forecast_date]
            actual_value = None
            if not matching.empty and model_type in matching.columns:
                actual = pd.to_numeric(matching[model_type], errors="coerce").dropna()
                if not actual.empty:
                    actual_value = round(float(actual.sum()), 2)
            forecast_points.append({
                "date": str(forecast_date), "value": round(predicted_value, 2),
                "actual": actual_value,
            })

        predicted_value = forecast_points[0]["value"]

        # Give the UI a meaningful, per-metric reference point. This makes
        # advice compare this forecast with the user's recent actual data.
        recent_baseline = None
        if model_type in df.columns:
            recent_values = pd.to_numeric(df[model_type], errors="coerce").dropna().tail(7)
            if not recent_values.empty:
                recent_baseline = round(float(recent_values.mean()), 2)

        # Save the prediction
        file_id = dataset_id

        prediction_ids = []
        for point in forecast_points:
            actual_value = point["actual"]
            prediction_error = round(actual_value - point["value"], 2) if actual_value is not None else None
            prediction_id = run_query(
                """INSERT INTO predictions
                   (user_id, company_id, uploaded_file_id, prediction_type, prediction_date,
                    actual_value, predicted_value, prediction_error, model_name)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (user_id, company_id, file_id, model_type, point["date"], actual_value,
                 point["value"], prediction_error, prediction_info.get("model_name", "linear_regression")),
                commit=True,
            )["last_id"]
            prediction_ids.append(prediction_id)

        add_history(
            user_id, company_id, "prediction",
            f"{model_type.title()} forecast from {target_date}",
            prediction_id=prediction_ids[0], file_id=file_id,
            status="ok", details=f"periods={forecast_periods}; first_value={predicted_value}",
        )

        try:
            _generate_powerbi_resources(user_id, company_id, session["company_name"])
        except Exception as exc:
            app.logger.warning("Power BI resource refresh skipped after prediction: %s", exc)

        # Build the historical + prediction chart
        # A chart is an enhancement, never a reason for a valid ML result to
        # fail (some matplotlib backends can be unavailable on Windows).
        try:
            chart = _build_chart(df, model_type, forecast_points)
        except Exception as chart_exc:
            print(f"[finsight] Chart rendering skipped: {chart_exc}")
            chart = ""

        return jsonify({
            "success": True,
            "prediction": predicted_value,
            "date": str(target_date),
            "model_type": model_type,
            "forecast_periods": forecast_periods,
            "forecasts": forecast_points,
            "history": [
                {"date": str(pd.to_datetime(row["tx_date"]).date()), "value": float(row[model_type])}
                for _, row in df.dropna(subset=["tx_date", model_type]).tail(60).iterrows()
            ],
            "chart": chart,
            "prediction_id": prediction_ids[0],
            "prediction_ids": prediction_ids,
            "baseline": recent_baseline,
            "model": {
                "name": prediction_info.get("model_name", "linear_regression"),
                "estimated_error": float(prediction_info.get("estimated_error") or 0),
                "metrics": prediction_info.get("metrics") or {},
            },
        })
    except mysql.connector.Error:
        app.logger.exception("Database failure during prediction for user %s", user_id)
        return jsonify({"error": "Prediction could not be saved because the database is unavailable. Please try again later."}), 503
    except Exception:
        app.logger.exception("Prediction request failed for user %s", user_id)
        return jsonify({"error": "Prediction could not be completed. Please try again."}), 500


def _build_chart(df: pd.DataFrame, model_type: str, forecast_points: list[dict],
                 *, date_column: str = "tx_date") -> str:
    col = str(model_type)
    if col not in df.columns or date_column not in df.columns:
        return ""

    df_plot = df.copy()
    df_plot[date_column] = pd.to_datetime(df_plot[date_column], errors="coerce")
    df_plot[col] = pd.to_numeric(df_plot[col], errors="coerce")
    df_plot = df_plot.dropna(subset=[date_column, col]).sort_values(date_column)
    if df_plot.empty:
        return ""

    plt.figure(figsize=(10, 5))
    plt.plot(df_plot[date_column], df_plot[col], marker="o",
             linewidth=2, label=f"Historical {col.title()}")
    forecast_dates = [pd.to_datetime(point["date"]) for point in forecast_points]
    forecast_values = [point["value"] for point in forecast_points]
    plt.plot(forecast_dates, forecast_values, color="red", marker="o",
             linewidth=2, linestyle="--", zorder=5, label="Forecast")
    plt.title(f"{col.title()} Forecast")
    plt.xlabel("Date")
    plt.ylabel(col.title())
    plt.legend()
    plt.grid(alpha=0.3)
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=100)
    plt.close()
    buf.seek(0)
    return base64.b64encode(buf.getvalue()).decode()


@app.post("/risk-classify")
@login_required
def risk_classify():
    """Classify one date using only the current user's active upload."""
    user_id, company_id = _session_user_id(), _session_company_id()
    payload = request.get_json(silent=True) or {}
    try:
        selected_date = pd.to_datetime(payload.get("date")).date()
    except Exception:
        return jsonify({"error": _message("api.error.risk_invalid_date")}), 400
    try:
        dataset_id, df = _active_dataset_and_frame(user_id)
        risk_info = get_models(user_id).get("risk")
        if dataset_id is None or df is None or df.empty:
            return jsonify({"error": _message("api.error.risk_no_data")}), 400
        required_features = ("revenue", "expenses", "profit", "amount")
        if any(column not in df.columns for column in required_features):
            return jsonify({
                "error": "Risk classification requires actual revenue, expenses, profit, and amount columns in the active dataset."
            }), 400
        if "tx_date" not in df.columns or pd.to_datetime(df["tx_date"], errors="coerce").notna().sum() == 0:
            return jsonify({
                "error": "Risk classification requires a usable date column in the active dataset."
            }), 400
        if risk_info is None:
            training_df = df.copy()
            training_df["tx_date"] = pd.to_datetime(training_df["tx_date"], errors="coerce")
            training_df = training_df.dropna(subset=["tx_date"])
            training_df["Date_Number"] = (training_df["tx_date"] - training_df["tx_date"].min()).dt.days
            training_df["Month"] = training_df["tx_date"].dt.month
            training_df["Day_of_Week"] = training_df["tx_date"].dt.dayofweek
            _train_models_for(user_id, training_df)
            risk_info = get_models(user_id).get("risk")
        if risk_info is None:
            return jsonify({"error": get_models(user_id).get(
                "risk_error", "Risk classifier training did not produce a model for this dataset."
            )}), 400
        matching = df[pd.to_datetime(df["tx_date"], errors="coerce").dt.date == selected_date]
        features = risk_info["features"]
        is_forecast = matching.empty
        if is_forecast:
            models = get_models(user_id)
            values = {}
            for column in features:
                model_info = models.get(column)
                if model_info is None:
                    return jsonify({"error": "Forecast models are unavailable for this dataset. Upload more varied dated data first."}), 400
                if get_pipeline is not None:
                    value, prediction_info = get_pipeline().predict_regression(
                        user_id, column, {}, history_df=df, prediction_date=selected_date
                    )
                    if value is None:
                        return jsonify({"error": prediction_info.get("error", "Risk forecast input is invalid.")}), 400
                    values[column] = float(value)
                else:
                    date_number = (pd.to_datetime(selected_date) - model_info["df_min"]).days
                    future_features = pd.DataFrame({"Date_Number": [date_number], "Month": [selected_date.month],
                                                    "Day_of_Week": [selected_date.weekday()]})
                    values[column] = float(model_info["model"].predict(future_features)[0])
        else:
            actual_features = matching[list(features)].apply(pd.to_numeric, errors="coerce")
            if actual_features.isna().any().any():
                return jsonify({
                    "error": "Risk classification cannot be completed because the selected date has missing or invalid required values."
                }), 400
            values = {column: float(actual_features[column].sum()) for column in features}
        if not all(np.isfinite(float(value)) for value in values.values()):
            return jsonify({"error": "Risk classification input contains missing or non-finite values."}), 400
        if get_pipeline is not None:
            risk_level, risk_prediction = get_pipeline().predict_risk(user_id, values)
            if risk_level is None:
                return jsonify({"error": risk_prediction.get("error", "Risk prediction input is invalid.")}), 400
        else:
            risk_input = pd.DataFrame([values], columns=features).apply(pd.to_numeric, errors="coerce")
            if not np.isfinite(risk_input.to_numpy()).all():
                return jsonify({"error": "Risk prediction input contains missing or non-finite values."}), 400
            risk_level = str(risk_info["model"].predict(risk_input)[0])
        if values["profit"] < 0:
            explanation = _message("risk.explain.negative_forecast" if is_forecast else "risk.explain.negative_actual")
        elif values["revenue"] > 0 and values["expenses"] > values["revenue"]:
            explanation = _message("risk.explain.expense_over_forecast" if is_forecast else "risk.explain.expense_over_actual")
        elif values["revenue"] > 0 and values["profit"] / values["revenue"] < 0.08:
            explanation = _message("risk.explain.margin_low_forecast" if is_forecast else "risk.explain.margin_low_actual")
        else:
            explanation = _message("risk.explain.stable_forecast" if is_forecast else "risk.explain.stable_actual")
        risk_id = run_query(
            """INSERT INTO risk_classifications
               (user_id, company_id, uploaded_file_id, classification_date, risk_level,
                revenue, expenses, profit, amount, explanation)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (user_id, company_id, dataset_id, selected_date, risk_level, values["revenue"],
             values["expenses"], values["profit"], values["amount"], explanation), commit=True,
        )["last_id"]
        add_history(user_id, company_id, "risk_classification", f"{risk_level} for {selected_date}",
                    file_id=dataset_id, status="ok", details=explanation)
        return jsonify({"success": True, "risk_id": risk_id, "risk_level": risk_level,
                        "date": str(selected_date), "revenue": values["revenue"],
                        "expenses": values["expenses"], "profit": values["profit"],
                        "amount": values["amount"], "explanation": explanation})
    except mysql.connector.Error:
        return jsonify({"error": "Risk classification could not be saved because the database is unavailable."}), 503
    except Exception as exc:
        print(f"[finsight] Risk classification failed: {exc}")
        return jsonify({"error": "Risk classification could not be completed."}), 500


# =====================================================
# Routes - History
# =====================================================
@app.route("/history")
@login_required
def history():
    user_id = _session_user_id()
    company_id = _session_company_id()
    events = run_query(
        """
        SELECT
            h.id, h.event_type, h.event_title, h.status, h.details, h.created_at,
            h.file_id, h.prediction_id,
            u.original_name  AS file_name,
            p.prediction_type, p.prediction_date, p.predicted_value
        FROM dashboard_history h
        LEFT JOIN uploaded_files u ON u.id = h.file_id
        LEFT JOIN predictions    p ON p.prediction_id = h.prediction_id
        WHERE h.user_id = %s
        ORDER BY h.created_at DESC
        LIMIT 200
        """,
        (user_id,),
        fetchall=True,
    )["rows"] or []

    return render_template(
        "history.html",
        company_name=session["company_name"],
        history=events,
    )


@app.get("/database")
@login_required
def database():
    """Compatibility view for the database-backed upload/prediction activity."""
    return history()


@app.get("/settings")
@login_required
def settings():
    """Show the signed-in account and active application settings."""
    user_columns = _load_user_columns()
    company_column = user_columns.get("company_id")
    email_column = user_columns.get("email")
    created_column = user_columns.get("created_at")
    if company_column:
        company_sql = f"u.{_quote_identifier(company_column)} AS company_id, c.company_name"
        company_join = "JOIN companies c ON c.id = u." + _quote_identifier(company_column)
    else:
        company_sql = "NULL AS company_id, '' AS company_name"
        company_join = ""
    email_sql = (
        f"u.{_quote_identifier(email_column)} AS email"
        if email_column else "'' AS email"
    )
    created_sql = (
        f"u.{_quote_identifier(created_column)} AS created_at"
        if created_column else "NULL AS created_at"
    )
    account = run_query(
        f"""SELECT {_user_display_expression(user_columns)},
                  {email_sql}, {created_sql},
                  {company_sql}
           FROM users u {company_join}
           WHERE u.id=%s""",
        (_session_user_id(),), fetchone=True,
    )["row"] or {
        "name": session.get("user_name", ""),
        "email": "",
        "company_name": session.get("company_name", ""),
        "created_at": None,
    }
    return render_template(
        "settings.html",
        company_name=session["company_name"],
        account=account,
        supported_formats="CSV, XLSX, XLS, JSON",
        max_upload_mb=round(MAX_CONTENT_LENGTH / (1024 * 1024), 2),
    )


@app.route("/api/history")
@login_required
def api_history():
    user_id = _session_user_id()
    events = run_query(
        """
        SELECT id, event_type, event_title, status, details, created_at
        FROM dashboard_history
        WHERE user_id = %s
        ORDER BY created_at DESC
        """,
        (user_id,),
        fetchall=True,
    )["rows"] or []

    for e in events:
        e["created_at"] = e["created_at"].isoformat() if e.get("created_at") else None
    return jsonify(events)


# =====================================================
# Routes - Power BI
# =====================================================
@app.route("/powerbi")
@login_required
def powerbi():
    """Local Power BI Desktop exports for the signed-in user's active upload."""
    user_id = _session_user_id()
    dataset_id = _current_dataset_id(user_id)
    return render_template(
        "powerbi.html",
        company_name=session["company_name"],
        has_data=dataset_id is not None,
    )


def _powerbi_export_frame(frame: pd.DataFrame, *, source: bool = False) -> pd.DataFrame:
    """Return a copy of an uploaded/derived frame without inventing fields."""
    work = frame.copy() if frame is not None else pd.DataFrame()
    if work.empty and len(work.columns) == 0:
        return work
    if source:
        detection = detect_columns(work)
        mapping = detection.get("mapping", {})
        try:
            work, _ = clean_dataframe(work, mapping)
        except ValueError:
            work = apply_mapping(work, mapping)
    work.columns = [str(column).strip() for column in work.columns]
    work = work.loc[:, ~work.columns.duplicated(keep="first")]
    # Excel cells cannot contain native JSON arrays/objects. Preserve those
    # uploaded values as JSON text for exports rather than dropping or
    # replacing them with placeholders.
    for column in work.columns:
        work[column] = work[column].map(
            lambda value: json.dumps(value, ensure_ascii=False, default=str)
            if isinstance(value, (list, dict)) else value
        )
    return work.reset_index(drop=True)


def _excel_download(frame: pd.DataFrame, filename: str):
    """Return one in-memory .xlsx file without exposing a server file path."""
    frame = _powerbi_export_frame(frame)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        frame.to_excel(writer, sheet_name="Data", index=False)
        worksheet = writer.sheets["Data"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells[:200]) + 2
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(width, 35)
    output.seek(0)
    return send_file(
        output, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _write_export_sheet(writer, sheet_name: str, frame: pd.DataFrame) -> None:
    """Write one dynamic export sheet while preserving its actual columns."""
    frame = frame if isinstance(frame, pd.DataFrame) else pd.DataFrame()
    frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    worksheet = writer.sheets[sheet_name[:31]]
    if worksheet.max_column:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            width = max(len(str(cell.value or "")) for cell in list(column_cells)[:200]) + 2
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(width, 35)


def _dynamic_dashboard_download(user_id: int, company_name: str, dataset_id: int):
    """Build a workbook dashboard from only the active upload's real data."""
    frames = _powerbi_export_frames(user_id, company_name, dataset_id)
    cleaned = frames.get("Cleaned_Data", pd.DataFrame())
    if cleaned.empty:
        return None

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        dashboard = writer.book.create_sheet("Dashboard", 0)
        dashboard.sheet_view.showGridLines = False
        dashboard["A1"] = f"FinSight AI | {company_name} | Active dataset"
        dashboard["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        dashboard["A1"].fill = PatternFill("solid", fgColor="17324D")
        dashboard.merge_cells("A1:H1")
        dashboard["A2"] = "Generated from the uploaded dataset."
        dashboard["A2"].font = Font(italic=True, color="5B6B7A")
        dashboard.merge_cells("A2:H2")

        kpi_frame = frames.get("KPI_Summary", pd.DataFrame())
        if not kpi_frame.empty:
            values = kpi_frame.iloc[0].to_dict()
            row = 4
            for name, value in values.items():
                if name in {"Company", "Exported_At", "Predictions"} or value is None:
                    continue
                dashboard.cell(row=row, column=1).value = str(name).replace("_", " ").title()
                dashboard.cell(row=row, column=2).value = value
                row += 1

        for sheet_name, frame in frames.items():
            if sheet_name in {"README", "Predictions", "Prediction_vs_Actual"}:
                continue
            _write_export_sheet(writer, sheet_name, frame)

        time_frame = frames.get("Time_Analysis", pd.DataFrame())
        if not time_frame.empty and len(time_frame.columns) > 1:
            sheet = writer.book["Time_Analysis"]
            chart = LineChart()
            chart.title = "Uploaded data over time"
            chart.y_axis.title = "Value"
            chart.x_axis.title = str(time_frame.columns[0])
            chart.height, chart.width = 8, 15
            chart.add_data(Reference(sheet, min_col=2, max_col=len(time_frame.columns),
                                     min_row=1, max_row=len(time_frame) + 1), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=len(time_frame) + 1))
            dashboard.add_chart(chart, "D4")

        category_frame = frames.get("Category_Analysis", pd.DataFrame())
        if not category_frame.empty and len(category_frame.columns) > 1:
            sheet = writer.book["Category_Analysis"]
            chart = BarChart()
            chart.type = "bar"
            chart.title = "Uploaded data by category"
            chart.height, chart.width = 8, 15
            chart.add_data(Reference(sheet, min_col=2, max_col=2, min_row=1,
                                     max_row=len(category_frame) + 1), titles_from_data=True)
            chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=len(category_frame) + 1))
            dashboard.add_chart(chart, "D20")

        for column in range(1, 15):
            dashboard.column_dimensions[chr(64 + column)].width = 16

    output.seek(0)
    return send_file(
        output, as_attachment=True, download_name="finsight_visual_dashboard.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _active_upload_record(user_id: int) -> dict | None:
    return run_query(
        """SELECT id, original_name, stored_name, raw_data FROM uploaded_files
           WHERE user_id=%s AND status='processed' ORDER BY id DESC LIMIT 1""",
        (user_id,), fetchone=True,
    )["row"]


def _upload_mimetype(filename: str) -> str:
    return mimetypes.guess_type(filename)[0] or "application/octet-stream"


def _raw_upload_frame(upload: dict) -> pd.DataFrame:
    """Read the exact uploaded bytes without applying business cleaning."""
    raw_data = upload.get("raw_data")
    if raw_data is not None:
        source = io.BytesIO(bytes(raw_data))
    else:
        resource = _powerbi_resource(_session_user_id(), create=False)
        source = (
            POWERBI_ROOT / resource["folder_token"] / "uploads" / upload["stored_name"]
            if resource else None
        )
        if source is None or not source.is_file():
            raise ValueError("The original uploaded file is unavailable.")
    return _read_upload_dataframe(source, upload["original_name"])


@app.route("/powerbi/download-excel/<export_type>")
@login_required
def download_powerbi_excel(export_type: str):
    """Download raw, cleaned, or dashboard data for only the active upload."""
    if export_type not in {"raw", "cleaned", "dashboard"}:
        return jsonify({"error": "Unknown Power BI export."}), 404

    user_id = _session_user_id()
    upload = _active_upload_record(user_id)
    if not upload:
        return jsonify({"error": "No processed dataset has been uploaded for this user."}), 404
    dataset_id = int(upload["id"])

    try:
        if export_type == "dashboard":
            dashboard = _dynamic_dashboard_download(user_id, session["company_name"], dataset_id)
            if dashboard is None:
                return jsonify({"error": "Cleaned data does not exist for the current dataset."}), 404
            return dashboard

        if export_type == "raw":
            # Return the exact bytes supplied by the user so the raw download
            # cannot accidentally become the cleaned or canonical projection.
            if upload.get("raw_data") is not None:
                return send_file(io.BytesIO(bytes(upload["raw_data"])),
                                 as_attachment=True,
                                 download_name=upload["original_name"],
                                 mimetype=_upload_mimetype(upload["original_name"]))
            resource = _powerbi_resource(user_id, create=False)
            target = (POWERBI_ROOT / resource["folder_token"] / "uploads" / upload["stored_name"]
                      if resource else None)
            if target is None or not target.is_file():
                return jsonify({"error": "The original uploaded file is unavailable."}), 404
            return send_file(target, as_attachment=True,
                             download_name=upload["original_name"],
                             mimetype=_upload_mimetype(upload["original_name"]))

        if export_type == "cleaned":
            frames = _powerbi_export_frames(user_id, session["company_name"], dataset_id)
            cleaned = frames["Cleaned_Data"].copy()
            if cleaned.empty:
                return jsonify({"error": "Cleaned data does not exist for the current dataset."}), 404
            return _excel_download(cleaned, "finsight_cleaned_data.xlsx")

        return jsonify({"error": "This Excel export is no longer available. Use the dashboard or prediction page."}), 404
    except mysql.connector.Error:
        app.logger.exception("Database failure during Power BI %s export", export_type)
        return jsonify({"error": "The export could not be created because the database is unavailable. Please try again later."}), 503
    except (TypeError, OSError, ImportError, UnicodeError, ValueError,
            pd.errors.ParserError) as exc:
        print(f"[finsight] Power BI {export_type} export failed: {exc}")
        return jsonify({"error": f"Could not create the {export_type} export. Please try again."}), 500


def _export_dimension_text(value: object) -> str:
    """Render a categorical value safely, including nested JSON values."""
    if value is None:
        return "Unspecified"
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    try:
        missing = pd.isna(value)
        if isinstance(missing, (bool, np.bool_)) and missing:
            return "Unspecified"
    except (TypeError, ValueError):
        pass
    return str(value)


def _reclean_uploaded_frame(user_id: int, dataset_id: int) -> pd.DataFrame | None:
    """Reload the exact upload and rerun the same cleaner for exports.

    Current uploads keep both the original bytes and the selected mapping in
    MySQL. Re-cleaning here makes the cleaned download reflect the pipeline at
    download time instead of merely echoing a previously generated artifact.
    Legacy records without raw bytes use their persisted cleaned rows below.
    """
    upload = run_query(
        """SELECT original_name, raw_data, column_mapping
           FROM uploaded_files WHERE id=%s AND user_id=%s AND status='processed'""",
        (dataset_id, user_id), fetchone=True,
    )["row"]
    if not upload:
        raise ValueError("The selected dataset is no longer available.")
    if upload.get("raw_data") is None:
        return None
    try:
        raw_frame = _read_upload_dataframe(
            io.BytesIO(bytes(upload["raw_data"])), upload["original_name"]
        )
        mapping = json.loads(upload.get("column_mapping") or "{}")
        if not isinstance(mapping, dict):
            mapping = {}
        cleaned, _ = clean_dataframe(raw_frame, mapping)
        return cleaned
    except (TypeError, ValueError, OSError, ImportError, UnicodeError, pd.errors.ParserError) as exc:
        raise ValueError(f"The uploaded file could not be cleaned again: {exc}") from exc


def _powerbi_export_frames(user_id: int, company_name: str,
                           dataset_id: int | None = None) -> dict[str, pd.DataFrame]:
    """Build export tables from this upload's preserved cleaned rows only."""
    if dataset_id is None:
        dataset_id = _current_dataset_id(user_id)
    if dataset_id is None:
        return {
            "Cleaned_Data": pd.DataFrame(), "KPI_Summary": pd.DataFrame(),
            "Time_Analysis": pd.DataFrame(), "Category_Analysis": pd.DataFrame(),
            "README": pd.DataFrame(),
        }
    cleaned = _reclean_uploaded_frame(user_id, dataset_id)
    if cleaned is None:
        stored_frame = _dataset_rows_frame(user_id, dataset_id)
    else:
        stored_frame = cleaned
    if stored_frame.empty:
        # Legacy uploads pre-dating dataset_rows can still be exported from
        # their real compatibility rows. A database failure is allowed to
        # propagate to the route's explicit database error response.
        cleaned_rows = run_query(
            """SELECT tx_date, transaction_id, description, amount, revenue, expenses, profit,
                      customers, marketing_spend, tx_type, category, payment_method,
                      department, city, status
               FROM financial_data WHERE user_id=%s AND uploaded_file_id=%s ORDER BY tx_date, id""",
            (user_id, dataset_id), fetchall=True,
        )["rows"] or []
        cleaned = pd.DataFrame(cleaned_rows)
    else:
        cleaned = stored_frame.copy()
    cleaned = _powerbi_export_frame(cleaned)

    types = _analysis_types_for(cleaned, use_cached=False)
    numeric_columns = [column for column, kind in types.items()
                       if kind == "numeric" and column in cleaned.columns
                       and pd.to_numeric(cleaned[column], errors="coerce").notna().any()
                       and not _visual_identifier_like(str(column), cleaned[column])]
    for column in numeric_columns:
        cleaned[column] = pd.to_numeric(cleaned[column], errors="coerce")
    date_column = _date_column_for(cleaned, types)
    visual = _visualization_data(cleaned, types)
    primary_metric = visual.get("metric_column")
    if primary_metric in numeric_columns:
        numeric_columns = [primary_metric] + [column for column in numeric_columns if column != primary_metric]

    # Predictions remain available to the private Power BI CSV integration and
    # the in-app history. They are deliberately excluded from the downloadable
    # Excel dashboard and from the removed predictions Excel endpoint.
    prediction_rows = run_query(
        """SELECT prediction_id AS Prediction_ID, prediction_type AS Prediction_Type,
                  prediction_date AS Prediction_Date, actual_value AS Actual_Value,
                  predicted_value AS Predicted_Value, prediction_error AS Prediction_Error,
                  model_name AS Model, created_at AS Created_At
           FROM predictions WHERE user_id=%s AND uploaded_file_id=%s
           ORDER BY prediction_date, prediction_id""",
        (user_id, dataset_id), fetchall=True,
    )["rows"] or []
    predictions = _powerbi_export_frame(pd.DataFrame(prediction_rows))

    kpi_values: dict[str, object] = {
        "Company": company_name,
        "Exported_At": datetime.now().replace(microsecond=0),
        "Cleaned_Rows": int(len(cleaned)),
        "Predictions": int(len(predictions)),
    }
    for column in numeric_columns:
        values = pd.to_numeric(cleaned[column], errors="coerce")
        prefix = "Total" if _visual_additive_measure(str(column)) else "Average"
        kpi_values[f"{prefix}_{str(column).replace(' ', '_')}"] = float(
            values.sum() if prefix == "Total" else values.mean()
        )
    if date_column and not cleaned.empty:
        dates = pd.to_datetime(cleaned[date_column], errors="coerce").dropna()
        if not dates.empty:
            kpi_values["First_Date"] = dates.min()
            kpi_values["Last_Date"] = dates.max()
    kpis = pd.DataFrame([kpi_values])

    period_label = "Period" if "Period" not in numeric_columns else "Date_Period"
    row_count_label = "Rows" if "Rows" not in numeric_columns else "Row_Count"
    time_analysis = pd.DataFrame(columns=[period_label, *numeric_columns, row_count_label])
    if date_column and numeric_columns and not cleaned.empty:
        dated = cleaned.copy()
        dated[date_column] = pd.to_datetime(dated[date_column], errors="coerce")
        dated = dated.dropna(subset=[date_column])
        if not dated.empty:
            internal_period = "__finsight_period__"
            while internal_period in dated.columns:
                internal_period += "_"
            dated[internal_period] = dated[date_column].dt.to_period("M").astype(str)
            time_analysis = dated.groupby(internal_period, as_index=False).agg(
                **{column: (column, "sum" if _visual_additive_measure(str(column)) else "mean")
                   for column in numeric_columns},
                **{row_count_label: (date_column, "size")},
            )
            time_analysis = time_analysis.rename(columns={internal_period: period_label})
            time_analysis = time_analysis[[period_label, *numeric_columns, row_count_label]]

    dimensions = [column for column, kind in types.items()
                  if kind in {"categorical", "text", "boolean"}
                  and column in cleaned.columns and cleaned[column].notna().any()]
    category_analysis = pd.DataFrame(columns=["Category", "Value"])
    if dimensions:
        dimension = dimensions[0]
        grouped = cleaned[[dimension] + ([numeric_columns[0]] if numeric_columns else [])].copy()
        grouped[dimension] = grouped[dimension].map(_export_dimension_text)
        if numeric_columns:
            primary = numeric_columns[0]
            grouped = grouped.groupby(dimension, as_index=False)[primary].agg(
                "sum" if _visual_additive_measure(str(primary)) else "mean"
            )
            category_analysis = grouped.rename(columns={dimension: "Category", primary: "Value"})
        else:
            category_analysis = grouped.groupby(dimension, as_index=False).size()
            category_analysis = category_analysis.rename(columns={dimension: "Category", "size": "Value"})
        category_analysis = category_analysis.sort_values("Value", ascending=False).head(50)

    distributions: dict[str, pd.DataFrame] = {}
    used_sheet_names = {
        "Cleaned_Data", "Predictions", "Prediction_vs_Actual", "KPI_Summary",
        "Time_Analysis", "Category_Analysis", "README",
    }
    for dimension in dimensions[:5]:
        grouped = cleaned[[dimension] + numeric_columns].copy()
        grouped[dimension] = grouped[dimension].map(_export_dimension_text)
        row_count_label = "Rows" if "Rows" not in numeric_columns else "Row_Count"
        aggregate = grouped.groupby(dimension, as_index=False).size().rename(columns={"size": row_count_label})
        if numeric_columns:
            totals = grouped.groupby(dimension, as_index=False)[numeric_columns].sum(min_count=1)
            aggregate = aggregate.merge(totals, on=dimension, how="left")
        sheet_key = re.sub(r"[^A-Za-z0-9]+", "_", str(dimension)).strip("_")[:22] or "Dimension"
        candidate = f"{sheet_key}_Analysis"
        suffix = 2
        while candidate in used_sheet_names:
            candidate = f"{sheet_key[:max(1, 22 - len(str(suffix)) - 1)]}_{suffix}_Analysis"
            suffix += 1
        used_sheet_names.add(candidate)
        distributions[candidate] = aggregate.sort_values(row_count_label, ascending=False).head(100)

    prediction_comparison = pd.DataFrame(
        columns=["Prediction_ID", "Prediction_Type", "Prediction_Date", "Actual_Value", "Predicted_Value", "Difference"]
    )
    if not predictions.empty and date_column and date_column in cleaned.columns:
        actual_dates = pd.to_datetime(cleaned[date_column], errors="coerce").dt.date
        comparison_rows = []
        for _, prediction in predictions.iterrows():
            target = str(prediction.get("Prediction_Type") or "")
            prediction_date = pd.to_datetime(prediction.get("Prediction_Date"), errors="coerce")
            actual = None
            if target in cleaned.columns and not pd.isna(prediction_date):
                matching = cleaned[actual_dates == prediction_date.date()]
                if not matching.empty:
                    values = pd.to_numeric(matching[target], errors="coerce").dropna()
                    if not values.empty:
                        actual = float(values.sum())
            predicted = float(prediction["Predicted_Value"])
            comparison_rows.append({
                "Prediction_ID": prediction["Prediction_ID"],
                "Prediction_Type": target,
                "Prediction_Date": prediction_date.date() if not pd.isna(prediction_date) else None,
                "Actual_Value": actual,
                "Predicted_Value": predicted,
                "Difference": actual - predicted if actual is not None else None,
            })
        prediction_comparison = pd.DataFrame(comparison_rows)

    readme = pd.DataFrame([
        {"Item": "Source", "Value": "The active upload's persisted cleaned rows"},
        {"Item": "Columns", "Value": ", ".join(map(str, cleaned.columns))},
        {"Item": "Numeric measures", "Value": ", ".join(map(str, numeric_columns)) or "None detected"},
        {"Item": "Date column", "Value": date_column or "None detected"},
    ])
    return {
        "Cleaned_Data": cleaned, "Predictions": predictions,
        "Prediction_vs_Actual": prediction_comparison, "KPI_Summary": kpis,
        "Time_Analysis": time_analysis, "Category_Analysis": category_analysis,
        **distributions, "README": readme,
    }


def _generate_powerbi_resources(user_id: int, company_id: int, company_name: str,
                                dataset_id: int | None = None) -> dict:
    """Create/update private CSV sources and a copy of the local Desktop template.

    The browser never receives filesystem paths.  The resource directory uses an
    opaque token and is authorized by `user_id` at every download endpoint.
    """
    resource = _powerbi_resource(user_id)
    if dataset_id is None:
        dataset_id = _current_dataset_id(user_id)
    paths = _powerbi_paths(resource, user_id, dataset_id)
    paths["data"].mkdir(parents=True, exist_ok=True)
    frames = _powerbi_export_frames(user_id, company_name, dataset_id)
    for sheet_name, frame in frames.items():
        if sheet_name == "README":
            continue
        filename = sheet_name.lower() + ".csv"
        frame.to_csv(paths["data"] / filename, index=False, encoding="utf-8-sig")
    frames["Cleaned_Data"].to_csv(paths["financial"], index=False, encoding="utf-8-sig")
    frames["Predictions"].to_csv(paths["predictions"], index=False, encoding="utf-8-sig")
    # Always copy the neutral visual template into the active user's private
    # dataset folder. This prevents a stale PBIX from a previous dataset/template
    # version being returned to the current user.
    if POWERBI_TEMPLATE.is_file():
        shutil.copy2(POWERBI_TEMPLATE, paths["pbix"])
    paths["readme"].write_text(
        "FinSight AI Power BI Desktop\n\n"
        "This folder belongs to one authenticated FinSight user. In Power BI Desktop, "
        "set the existing Cleaned_Data query source to data/financial_data.csv and the "
        "Predictions query source to data/predictions.csv, then save and use Refresh. "
        "Additional visualization-ready CSV tables are in the same data folder. "
        "Do not connect this report to the shared MySQL database.\n\n"
        "The Flask app generates data and copies the reusable PBIX template; it does not "
        "edit or create a PBIX file.\n",
        encoding="utf-8",
    )
    run_query(
        """UPDATE user_powerbi_resources SET financial_csv=%s, predictions_csv=%s,
           pbix_filename=%s, generated_at=NOW() WHERE user_id=%s""",
        ("data/financial_data.csv", "data/predictions.csv", paths["pbix"].name if paths["pbix"].exists() else None, user_id),
        commit=True,
    )
    if dataset_id is not None and paths["pbix"].exists():
        run_query(
            """INSERT INTO powerbi_desktop_reports
               (user_id, uploaded_file_id, pbix_filename, status)
               VALUES (%s, %s, %s, 'generated')
               ON DUPLICATE KEY UPDATE pbix_filename=VALUES(pbix_filename), status='generated'""",
            (user_id, dataset_id, paths["pbix"].name), commit=True,
        )
    return {"resource": resource, "paths": paths}


@app.route("/powerbi/export")
@login_required
def export_powerbi_data():
    """Download one Excel workbook that Power BI Desktop can import locally."""
    user_id = _session_user_id()
    company_id = _session_company_id()
    company_name = session["company_name"]
    dataset_id = _current_dataset_id(user_id)
    if dataset_id is None:
        return jsonify({"error": "No processed dataset has been uploaded for this user."}), 404

    # Keep this endpoint on the same dataset-scoped, schema-agnostic builder as
    # the dashboard download. The old implementation assumed financial sheet
    # names and would fail or show empty financial KPIs for other datasets.
    try:
        response = _dynamic_dashboard_download(user_id, company_name, dataset_id)
    except (TypeError, ValueError, OSError, ImportError, UnicodeError, pd.errors.ParserError) as exc:
        app.logger.warning("Power BI workbook export could not be built: %s", exc)
        return jsonify({"error": "Could not create a workbook from the current dataset. Please try again."}), 500
    if response is None:
        return jsonify({"error": "Cleaned data does not exist for the current dataset."}), 404
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response.headers["Content-Disposition"] = (
        f'attachment; filename="FinSight_powerbi_data_{timestamp}.xlsx"'
    )
    add_history(user_id, company_id, "powerbi_export",
                f"FinSight_powerbi_data_{timestamp}.xlsx", status="ok",
                details="Power BI-ready workbook downloaded")
    return response


@app.post("/powerbi/generate")
@login_required
def generate_powerbi_data():
    user_id = _session_user_id()
    dataset_id = _current_dataset_id(user_id)
    if dataset_id is None:
        message = "Upload and process a dataset before generating Power BI resources."
        if request.is_json:
            return jsonify({"error": message}), 404
        flash(message, "warning")
        return redirect(url_for("powerbi"))
    result = _generate_powerbi_resources(user_id, _session_company_id(), session["company_name"], dataset_id)
    add_history(user_id, _session_company_id(), "powerbi_generate", "Power BI Desktop data updated", status="ok")
    if not request.is_json:
        return redirect(url_for("powerbi"))
    return jsonify({"success": True, "has_template": result["paths"]["pbix"].exists()})


@app.route("/powerbi/download/<resource_type>")
@login_required
def download_powerbi_resource(resource_type: str):
    user_id = _session_user_id()
    current_dataset_id = _current_dataset_id(user_id)
    if current_dataset_id is None:
        return jsonify({"error": "No processed dataset has been uploaded for this user."}), 404

    # Render's local filesystem is ephemeral. Rebuild the data resources from
    # the active upload on every request so a surviving database record can
    # never point the user at a stale or unrelated local CSV.
    try:
        _generate_powerbi_resources(
            user_id, _session_company_id(), session["company_name"], current_dataset_id
        )
    except (TypeError, ValueError, OSError, ImportError, UnicodeError,
            pd.errors.ParserError) as exc:
        app.logger.warning("Power BI resource generation failed: %s", exc)
        return jsonify({"error": "Power BI resources could not be generated from the current dataset."}), 500

    resource = _powerbi_resource(user_id, create=False)
    if not resource:
        return jsonify({"error": "Power BI resource not found"}), 404
    paths = _powerbi_paths(resource, user_id, current_dataset_id)
    allowed = {"template": paths["pbix"], "financial_data": paths["financial"],
               "predictions": paths["predictions"], "instructions": paths["readme"]}
    target = allowed.get(resource_type)
    if target is None or not target.is_file():
        return jsonify({"error": "Power BI resource not found"}), 404
    return send_file(target, as_attachment=True, download_name=target.name)


# =====================================================
# Authenticated upload download
# =====================================================
@app.route("/uploads/<int:file_id>")
@login_required
def uploaded_file(file_id: int):
    row = run_query("SELECT original_name, stored_name, raw_data FROM uploaded_files WHERE id=%s AND user_id=%s",
                    (file_id, _session_user_id()), fetchone=True)["row"]
    if not row:
        return jsonify({"error": "File not found"}), 404
    if row.get("raw_data") is not None:
        return send_file(
            io.BytesIO(bytes(row["raw_data"])), as_attachment=True,
            download_name=row["original_name"],
            mimetype=_upload_mimetype(row["original_name"]),
        )
    resource = _powerbi_resource(_session_user_id(), create=False)
    if not resource:
        return jsonify({"error": "File not found"}), 404
    target = POWERBI_ROOT / resource["folder_token"] / "uploads" / row["stored_name"]
    if not target.is_file():
        return jsonify({"error": "File not found"}), 404
    return send_file(target, as_attachment=True, download_name=row["original_name"],
                     mimetype=_upload_mimetype(row["original_name"]))


# =====================================================
# Error handlers
# =====================================================
@app.errorhandler(404)
def not_found(_):
    if request.path.startswith("/api/"):
        return jsonify({"error": "Not found"}), 404
    return render_template("404.html"), 404


@app.errorhandler(413)
def too_large(_):
    limit = round(MAX_CONTENT_LENGTH / (1024 * 1024), 2)
    return jsonify({"error": f"Uploaded file is too large. Maximum size is {limit:g} MB."}), 413


@app.errorhandler(RuntimeError)
def runtime_configuration_error(error):
    """Keep configuration/runtime failures useful without exposing internals."""
    app.logger.error("Application runtime failure: %s", error, exc_info=True)
    message = "The application is temporarily unavailable. Please try again later."
    if request.path.startswith("/api/") or request.is_json:
        return jsonify({"error": message}), 503
    flash(message, "danger")
    return render_template("500.html"), 503


@app.errorhandler(mysql.connector.Error)
def database_unavailable(error):
    """Return a useful response when MySQL is unavailable or misconfigured."""
    app.logger.error(
        "Database request failed: %s: %s",
        type(error).__name__,
        error,
        exc_info=(type(error), error, error.__traceback__),
    )
    message = "Database connection failed. Please try again later."
    if request.path.startswith("/api/") or request.is_json:
        return jsonify({"error": message}), 503
    if request.path == "/signup":
        flash(message, "danger")
        return render_template("signup.html"), 503
    if request.path == "/login":
        flash(message, "danger")
        return render_template("login.html"), 503
    return render_template("500.html"), 503


@app.errorhandler(500)
def server_error(_):
    if request.path.startswith("/api/"):
        return jsonify({"error": "Internal server error"}), 500
    return render_template("500.html"), 500


# =====================================================
# Entry point
# =====================================================
init_database()

if __name__ == "__main__":
    host = _env("FLASK_RUN_HOST", "0.0.0.0")
    port = _int_env("PORT", _int_env("FLASK_RUN_PORT", 5000))
    debug = _bool_env("FLASK_DEBUG", False)
    print(f"[finsight] Starting Flask on http://{host}:{port} (debug={debug})")
    app.run(host=host, port=port, debug=debug)
